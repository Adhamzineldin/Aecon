# Create your views here.
from django.http import HttpResponse, JsonResponse, Http404, HttpResponseRedirect, HttpResponseForbidden, \
    HttpResponseBadRequest, HttpResponseServerError
import itertools
import operator
import re
import time
from .scheduler import stop_timeout_job, start_timeout_job

from django.shortcuts import render
from django.template import loader
from django.conf import settings
import datetime
import requests
import json
from datetime import datetime
from django.db.models.functions import Cast
from aecon.models import *
import pandas as pd
import pytz
from dateutil import parser
from .forms import *
import os
from .automatedemails import sendMail, sendMail_withAttachments
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
import threading
import pymysql
from django.shortcuts import redirect
from django.contrib.auth import authenticate, login, logout
from . import automatedemails
import traceback
import openpyxl
from django.core.files.storage import FileSystemStorage
from django.contrib import messages
import warnings

from openpyxl import Workbook
from openpyxl.styles import Alignment
from tempfile import NamedTemporaryFile
import zipfile
import math
import time
from .utils import upload_jtc_data_todb, turning_count_thread
from django.utils.safestring import mark_safe

coords = [[]]

EXCEL_ROOT = settings.MEDIA_ROOT
BASE_URL = settings.BASE_URL
BASE_DIR = settings.BASE_DIR
LOGIN_URL = settings.LOGIN_URL
DB_NAME = settings.DB_NAME
SEND_MAIL_TO = settings.SEND_MAIL_TO
ALPHABET = "0123456789abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ"
os.makedirs(EXCEL_ROOT, exist_ok=True)


def generate_salt(length=128):
    return "".join(random.choice(ALPHABET) for i in range(length))


def calculate_num_datapoints(**kwargs):
    print(kwargs)
    numsites = len(kwargs["ids"])
    print("numsites is", numsites)
    numDays = (kwargs["endDate"] - kwargs["startDate"]).days
    print("num days is", numDays)
    numDirections = 2 if kwargs["direction"] == "split" else 1
    numClasses = 7  # len(kwargs["classes"])
    if "period" in kwargs:
        if kwargs["period"] == "D":
            numSegments = 24
        else:
            numSegments = 24 * (60 / int(kwargs["period"]))
    else:
        numSegments = 1
    return numsites * numDays * numDirections * numSegments * numClasses


def catch_database_errors(func):
    def decorator(request, *args, **kwargs):
        try:
            return func(request, *args, **kwargs)
        except django.db.utils.Error as e:
            automatedemails.sendMail(SEND_MAIL_TO, "school streets Server Error",
                                     "view_project , django error " + str(e))
            print("database error", e)
            return HttpResponseServerError("Database error. Please contact Tracsis")
        except pymysql.Error as e:
            automatedemails.sendMail(SEND_MAIL_TO, "school streets Server Error",
                                     "view_project , mysql error " + str(e))
            return HttpResponseServerError("MySQL error. Please contact Tracsis")

    return decorator


def must_be_admin(func):
    def decorator(request, *args, **kwargs):
        try:
            print("must be admin")
            assert request.user.is_authenticated and request.user.groups.filter(
                    name="admin").exists()
        except Exception as e:
            print("error was", e)
            return HttpResponseForbidden()
        return func(request, *args, **kwargs)

    return decorator


# def session_check(func):
#     def decorator(request, *args, **kwargs):
#         try:
#             session_creation = parser.parse(request.session['user'])
#             time_diff =  datetime.datetime.now() - session_creation
#             time_diff = time_diff.total_seconds()
#             time_diff = round(time_diff/60)
#             print("time diff -", time_diff)
#             if time_diff >= 60:
#                 del request.session['user']
#                 logout(request)
#             else:
#                 request.session['user'] = str(datetime.datetime.now())
#                 request.session.modified = True
#         except Exception as e:
#             request.session['user'] = str(datetime.datetime.now())
#         return func(request, *args, **kwargs)
#     return decorator

def get_admin_context(locations):
    context = {"observationtypes": ObservationType.objects.using(DB_NAME).all(),
               "locations":        Location.objects.using(DB_NAME).values("id", "name"), "form": LocationForm(),
               "sideBar":          locations.format_for_sidebar(style="crt"),
               "apis":             VivacityAPI.objects.using(DB_NAME).all(),
               "client":           Client.objects.using(DB_NAME).get(id=2)}
    return context


# Create your views here.


@login_required(login_url=LOGIN_URL)
@must_be_admin
def admin_all_locations(request):
    print("admin_location all test")
    template = loader.get_template(f'{BASE_URL}/all-locations-admin-view.html')
    clients = Client.objects.using(
            DB_NAME).all().exclude(id__in=[2, 4, 11, 15])
    # locs = Location.objects.using(DB_NAME).filter(client__in=clients, temp=0, observationType_id=1, api_identifier__isnull=False)
    locs = Location.objects.using(DB_NAME).filter(
            client__in=clients, temp=0, observationType_id=1)
    # process_old_metrolink_count_files()
    print(locs.count())
    locs = locs.get_admin_stats()

    # qs = clients[0].get_admin_stats()
    # for c in clients[1:]:
    #    qs = qs.union(c.get_admin_stats(), all=True)
    context = {"locations": locs, "clients": clients,
               "header":    "All Sensors Overview"}
    response = template.render(context, request)
    print(query_count_all(), "test")
    return HttpResponse(response)


@login_required(login_url=LOGIN_URL)
@must_be_admin
def admin_backfill(request):
    data = request.POST
    loc = Location.objects.using(DB_NAME).get(id=data["location_id"])

    date = datetime.datetime.strptime(data["date"], "%Y-%m-%d")
    print(loc, date)
    threading.Thread(target=run_admin_backfill,
                     args=(loc, date, request.user)).start()
    return JsonResponse({"status": "ok"})


def run_admin_backfill(loc, date, user):
    try:
        with transaction.atomic(using=DB_NAME):
            existingData = Observation.objects.using(DB_NAME).filter(location=loc, date__gte=date,
                                                                     date__lt=date + datetime.timedelta(days=1))
            existingTotals = DailyClassedTotals.objects.using(DB_NAME).filter(location=loc, date__gte=date,
                                                                              date__lt=date + datetime.timedelta(
                                                                                      days=1))  # .delete()
            print("num existing entries",
                  existingData.count(), existingTotals.count())
            existingData.delete()
            existingTotals.delete()
            api = VivacityAPI.objects.using(DB_NAME).prefetch_related(
                    "locations").get(locations=loc)
            api.backfill(date, date + datetime.timedelta(days=1),
                         selectedSites=[loc])
            print("finished backfilling")
            desc = "Site Backfilled for " + date.strftime("%d-%m-%Y")
            Event.objects.using(DB_NAME).create(location=loc, desc=desc, date=datetime.datetime.now().date(),
                                                icon="icon-edit-1", addedBy=user)
    except ValueError as e:
        print("error in admin backfill", e)


@login_required(login_url=LOGIN_URL)
@must_be_admin
def admin_all_locations_weekly(request):
    template = loader.get_template(
            f'{BASE_URL}/admin-overall-weekly-view.html')
    clients = Client.objects.using(DB_NAME).all().exclude(id__in=[2, 4, 11])
    result = list(clients[0].get_admin_weekly_stats())
    for c in clients[1:]:
        result += list(c.get_admin_weekly_stats())

    d = datetime.datetime.now().date()
    monday = d - datetime.timedelta(days=d.weekday())
    context = {"locations": result, "clients": clients, "d1": monday - datetime.timedelta(days=7),
               "d2":        monday - datetime.timedelta(days=14), "header": "All Sensors Overview - Weekly"}
    response = template.render(context, request)
    # print(query_count_all())
    return HttpResponse(response)


@login_required(login_url=LOGIN_URL)
@must_be_admin
def admin_events_view(request):
    template = loader.get_template(f'{BASE_URL}/admin-events-view.html')
    clients = Client.objects.using(DB_NAME).all().exclude(id__in=[2, 4])
    events = Event.objects.using(DB_NAME).all().order_by(
            "-date").prefetch_related("location")
    context = {"events":                  events, "client": Client.objects.using(
            DB_NAME).get(id=2), "header": "Events Overview"}
    return HttpResponse(template.render(context, request))


@login_required(login_url=LOGIN_URL)
@must_be_admin
def admin_location_view(request):
    if "id" in request.GET:
        location = Location.objects.using(DB_NAME).get(id=request.GET["id"])
    else:
        location = None
    template = loader.get_template(f'{BASE_URL}/location-view.html')
    sideBar = Project.objects.using(DB_NAME).all().format_for_sidebar()
    client = Client.objects.using(DB_NAME).get(id=2)
    context = {"client":           client, "DEBUG": settings.DEBUG,
               "loc":              location,
               "observationtypes": ObservationType.objects.using(DB_NAME).all(),
               "apis":             VivacityAPI.objects.using(DB_NAME).all(), "header": "Site Summary",
               "classes":          ObservationClass.objects.using(DB_NAME).all(),
               "events":           Event.objects.using(DB_NAME).filter(location=location).order_by("date").as_html()}
    print("context is", context)
    return HttpResponse(template.render(context, request))


@login_required(login_url=LOGIN_URL)
@require_POST
def get_view(request):
    view = request.POST["view"]
    request.session["current-view"] = view
    return JsonResponse({"redirect": f"/{BASE_URL}/admin"})


@login_required(login_url=LOGIN_URL)
def api_view(request):
    if request.method == "GET":
        data = request.GET.dict()
        if "id" in data:
            api = VivacityAPI.objects.using(DB_NAME).get(id=data["id"])
            print("api is", api.name)
            d = datetime.datetime.now().replace(microsecond=0) - datetime.timedelta(days=1)
            # api.get_data_from_api_and_save(d, d + datetime.timedelta(minutes=15), save=False)
            result = []
            print("result is", result)

        else:
            api = None
        template = loader.get_template(f'{BASE_URL}/api-view.html')
        context = {"client": Client.objects.using(DB_NAME).all(), "DEBUG": settings.DEBUG,
                   "apis":   VivacityAPI.objects.using(DB_NAME).all(),
                   "api":    api}
        print("context is", context)
        return HttpResponse(template.render(context, request))
    if request.method == "POST":
        data = request.POST
        print("data is", data)
        try:
            if "apiName" in data:
                api = VivacityAPI.objects.using(
                        DB_NAME).get(name=data["apiName"])
            else:
                api = VivacityAPI.create(name="api_test_object")
                if "APIKey" in data:
                    api.APIKey = data["APIKey"]
                if "baseUrl" in data:
                    api.baseUrl = data["baseUrl"]
            result = api.get_data_from_api(datetime.datetime.now() - datetime.timedelta(hours=1),
                                           datetime.datetime.now())
        except VivacityAPI.DoesNotExist as e:
            return JsonResponse({"status": "failed", "message": "API does not exist in database"})
        except requests.exceptions.RequestException as e:
            print("Requests exception", e)
            return JsonResponse({"status": "failed", "message": "unable to retrieve data, message was " + repr(e)})
        except json.decoder.JSONDecodeError as e:
            return JsonResponse({"status": "failed", "message": "Received response from API, but there was no data"})
        output = {"countlines": [], "classes": []}
        for r in result:
            for d, items in r.items():
                for item in items:
                    for cl in item["countlines"]:
                        if cl["countlineId"] not in output["countlines"]:
                            output["countlines"].append(cl["countlineId"])
                        for count in cl["counts"]:
                            if count["class"] not in output["classes"]:
                                output["classes"].append(count["class"])
        return JsonResponse({"status": "ok", "data": json.dumps(output, indent=4, sort_keys=True)})


@login_required(login_url=LOGIN_URL)
def get_location_data_daily_counts(request):
    try:
        selectedLocation = request.POST["location_id"]
        print("in get daily counts, location is", selectedLocation)
        year = request.POST["year"]
        print("year is", year)
        loc = Location.objects.using(DB_NAME).get(id=selectedLocation)
        result = loc.get_daily_data_point_counts(year)
        # print("result is", result)
        return JsonResponse({"result": result, "year": year})
    except Location.DoesNotExist as e:
        return HttpResponseBadRequest(reason="No such location")
    except django.db.utils.Error as e:
        automatedemails.sendMail(SEND_MAIL_TO, "school_streets Server Error",
                                 "get_location_data_daily_counts , django error " + str(e))
        return HttpResponseServerError
    except pymysql.Error as e:
        automatedemails.sendMail(SEND_MAIL_TO, "school_streets Server Error",
                                 "get_location_data_daily_counts , mysql error " + str(e))
        return HttpResponseServerError


@login_required(login_url=LOGIN_URL)
def get_location_data_hourly_counts(request):
    try:
        selectedLocation = request.POST["location_id"]
        date = request.POST["date"]

        loc = Location.objects.using(DB_NAME).get(id=selectedLocation)
        result = loc.get_daily_data_point_counts_hourly(date)
        return JsonResponse(result)
    except Location.DoesNotExist as e:
        return HttpResponseBadRequest(reason="No such location")
    except django.db.utils.Error as e:
        automatedemails.sendMail(SEND_MAIL_TO, "school_streets Server Error",
                                 "get_location_data_hourly_counts , django error " + str(e))
        return HttpResponseServerError
    except pymysql.Error as e:
        automatedemails.sendMail(SEND_MAIL_TO, "school_streets Server Error",
                                 "get_location_data_hourly_counts , mysql error " + str(e))
        return HttpResponseServerError
    except ValueError as e:
        return HttpResponseBadRequest(reason="Invalid Date Format")


@login_required(login_url=LOGIN_URL)
def get_location_directions(request):
    if request.method == "POST":
        return JsonResponse({"status": "ok", "data": Direction.objects.using(DB_NAME).all().as_html()})
    return JsonResponse({"status": "failed", "message": "incorrect request method"})


@login_required(login_url=LOGIN_URL)
def get_location_classes(request):
    if request.method == "POST":
        selectedLocation = request.POST["location_id"]
        loc = Location.objects.using(DB_NAME).get(id=selectedLocation)
        return JsonResponse({"status":         "ok",
                             "data":           loc.classes.all().as_list_of_selected_classes(),
                             "orderedClasses": ObservationClassSerializer(loc.ordered_class_list(), many=True).data,
                             "groups":         ObservationClassGroup.objects.using(DB_NAME).all().as_html_list()})
    return JsonResponse({"status": "failed", "message": "incorrect request method"})


@login_required(login_url=LOGIN_URL)
def save_location_classes(request):
    print("date is", request.POST)
    try:
        loc = Location.objects.using(DB_NAME).get(
                id=request.POST["location_id"])
        loc.save_classes(json.loads(request.POST["classes"]))
    except Location.DoesNotExist as e:
        return JsonResponse({"status": "failed", "message": "WEEEEEEEE"})
    return JsonResponse({"status":         "ok",
                         "data":           loc.classes.all().as_list_of_selected_classes(),
                         "orderedClasses": ObservationClassSerializer(loc.ordered_class_list(), many=True).data,
                         "groups":         ObservationClassGroup.objects.using(DB_NAME).all().as_html_list()})


@login_required(login_url=LOGIN_URL)
def apply_group(request):
    print("data is", request.POST)
    try:
        loc = Location.objects.using(DB_NAME).get(
                id=request.POST["location_id"])
        group = ObservationClassGroup.objects.using(
                DB_NAME).get(id=int(request.POST["group"]))
        print("selected group is", group)
        loc.save_classes([cl.id for cl in group.ordered_class_list()])
    except Location.DoesNotExist as e:
        return JsonResponse({"status": "failed", "message": "No such location in database"})
    except ObservationClassGroup.DoesNotExist as e:
        return JsonResponse({"status": "failed", "message": "No such group in database"})
    except ValueError as e:
        return JsonResponse({"status": "failed", "message": "Invalid group id"})
    return JsonResponse({"status":         "ok",
                         "data":           loc.classes.all().as_list_of_selected_classes(),
                         "orderedClasses": ObservationClassSerializer(loc.ordered_class_list(), many=True).data,
                         "groups":         ObservationClassGroup.objects.using(DB_NAME).all().as_html_list()})


@login_required(login_url=LOGIN_URL)
def get_api_list(request):
    if request.method == "POST":
        selectedLocation = request.POST["location_id"]
        return JsonResponse({"status": "ok", "data": Location.objects.using(DB_NAME).get(
                id=selectedLocation).classes.all().as_list_of_selected_classes()})
    return JsonResponse({"status": "failed", "message": "incorrect request method"})


@login_required(login_url=LOGIN_URL)
def save_location(request):
    if request.method == "POST":
        print("in save location , data is", request.POST)
        try:
            action = request.POST["action"]
            if action == "create":
                loc = Location.create(id=request.POST["locId"])
            elif action == "update":
                loc = Location.objects.using(
                        DB_NAME).get(id=request.POST["locId"])
            else:
                return JsonResponse({"status": "failed", "message": "Unrecognised action"})
            with transaction.atomic(using=DB_NAME):
                loc.update(request.POST, request.FILES)
            return JsonResponse({"status": "ok", "message": "Successfully saved location"})
        except KeyError as e:
            return JsonResponse({"status": "failed", "message": "unable to process data"})
        except django.db.IntegrityError as e:
            print(e)
            return JsonResponse({"status": "failed", "message": "Location with unique ID already exists in database"})
        except django.core.exceptions.ValidationError as e:
            for key, msg in e.message_dict.items():
                return JsonResponse({"status": "failed", "message": str(key) + ": " + msg[0]})
        except Location.DoesNotExist as e:
            return JsonResponse({"status": "failed", "message": "Location does not exist in database"})
        except Direction.DoesNotExist as e:
            return JsonResponse({"status": "failed", "message": "Selected Direction does not exist in database"})
    raise Http404


@catch_database_errors
@require_POST
@login_required(login_url=LOGIN_URL)
def get_locations(request):
    ids = json.loads(request.POST["location_id"])
    # set_cookie("Location", ids, max_age = None, expires = None)
    if type(ids) == list:
        loc = Location.objects.using(DB_NAME).filter(
                id__in=ids).order_by('name')
    else:
        loc = Location.objects.using(DB_NAME).filter(id=ids).order_by('name')

    if len(loc) == 0:
        return HttpResponseBadRequest(reason="No countlines specified")
    loc_ids = loc.values_list('id', flat=True)
    proj_loc = ProjectLocations.objects.using(DB_NAME).filter(
            location__in=loc_ids).order_by('startDate')
    project_lst = []
    for item in proj_loc:
        project = Project.objects.using(DB_NAME).get(id=item.project_id)
        if not item.speed_limit:
            project_lst.append(
                    {'id':       project.id, 'name': project.project_no, "start_date": item.startDate,
                     "end_date": item.endDate})
        else:
            project_lst.append(
                    {'id':         project.id, 'name': project.project_no, 'speed_limit': item.speed_limit,
                     "start_date": item.startDate,
                     "end_date":   item.endDate})
    try:
        area = loc[0].area.name
    except Exception as e:
        area = None

    arms_lst = []
    arms = Arms.objects.using(DB_NAME).filter(
            location__in=loc_ids).order_by('name')

    for arm in arms:
        arms_lst.append(
                {'id':      arm.id, 'name': arm.name, 'Fullname': arm.display_name, 'lat': arm.lat, 'lon': arm.lon,
                 'project': arm.project.id, 'location': arm.location.id})

    if loc[0].observationType_id in (11, 13):
        classes = []
        for cls in LocationObservationClass.objects.using(DB_NAME).filter(location=loc[0]):
            classes.append(
                    {'id': cls.id, 'name': cls.obsClass.displayName, 'order': cls.order})
    else:
        classes = loc[0].ordered_class_queryset().as_dashboard_class_list()

    return JsonResponse({"status":             "ok",
                         "locations":          loc.as_geojson(),
                         "classes":            classes,
                         "directions":         loc[0].ordered_direction_queryset(),
                         "startRecievingDate": loc[0].startRecievingDate,
                         "area":               area,
                         "name":               loc[0].name,
                         "installDate":        loc[0].installDate,
                         "lastDataReceived":   loc[0].lastDataReceived,
                         "sensor_status":      loc[0].status,
                         "project":            project_lst,
                         'arms':               arms_lst,
                         "validationDate":     loc[0].validationDate,
                         "imgURL":             loc[0].imgURL
                         })


@login_required(login_url=LOGIN_URL)
def upload_location_data(request):
    print(request.POST)
    loc = Location.objects.using(DB_NAME).get(id=request.POST["location_id"])
    if request.FILES:
        file_obj = request.FILES["data-file"]
        dir = os.path.join(settings.BASE_DIR, "data", BASE_URL)
        if not os.path.exists(dir):
            os.mkdir(dir)
        file_loc = os.path.join(dir, file_obj.name)
        with open(file_loc, 'wb+') as destination:
            for chunk in file_obj.chunks():
                destination.write(chunk)
        threading.Thread(target=loc.import_data_from_file, args=(
                file_loc, request.POST["startDate"])).start()
        return JsonResponse({"status": "failed", "message": "Upload of " + file_obj.name + " completed"})
    return JsonResponse({"status": "ok", "message": "Well, something happened anyway"})


@login_required(login_url=LOGIN_URL)
def get_API_classes(request):
    if request.method == "POST":
        return JsonResponse(
                {"status": "ok",
                 "data":   VivacityAPI.objects.using(DB_NAME).get(name="CRT").classes_as_html_menu_list()})
    return JsonResponse({"status": "failed", "message": "incorrect request method"})


@login_required(login_url=LOGIN_URL)
def save_vivacity_API(request):
    if request.method == "POST":
        try:
            action = request.POST["action"]
            if action == "create":
                api = VivacityAPI.create(name=request.POST["apiName"])
            elif action == "update":
                api = VivacityAPI.objects.using(DB_NAME).get(
                        name=request.POST["apiName"])
            else:
                return JsonResponse({"status": "failed", "message": "Unrecognised action"})
            api.update(request.POST)
            return JsonResponse({"status": "ok", "message": "Successfully saved new API"})
        except KeyError as e:
            print(e)
            return JsonResponse({"status": "failed", "message": "unable to process data"})
        except django.db.IntegrityError as e:
            return JsonResponse({"status": "failed", "message": "API with unique name already exists in database"})
        except django.core.exceptions.ValidationError as e:
            for key, msg in e.message_dict.items():
                return JsonResponse({"status": "failed", "message": str(key) + ": " + msg[0]})
        except VivacityAPI.DoesNotExist as e:
            return JsonResponse({"status": "failed", "message": "API does not exist in database"})
        except Direction.DoesNotExist as e:
            return JsonResponse({"status": "failed", "message": "Selected Direction does not exist in database"})
    raise Http404


@login_required(login_url=LOGIN_URL)
def ATC_view(request):
    loc = Location.objects.using(DB_NAME).get(id="dfhhjdj")
    client = request.user.client_set.all(
    ).prefetch_related("locations__area")[0]
    template = loader.get_template(f'{BASE_URL}/atc-view.html')
    response = template.render({"client": client, "location": loc}, request)
    return HttpResponse(response)


@login_required(login_url=LOGIN_URL)
def radar_view(request):
    loc = Location.objects.using(DB_NAME).get(id="dfhhjdj")
    client = request.user.client_set.all(
    ).prefetch_related("locations__area")[0]
    template = loader.get_template(f'{BASE_URL}/radar-view.html')
    response = template.render({"client": client, "location": loc}, request)
    return HttpResponse(response)


def borders_view(request):
    print("in borders view user is", request.user)
    if "borders" in request.get_full_path():
        client = Client.objects.using(DB_NAME).get(id=15)
        logo = "EL_Logo.JPG"
    else:
        client = Client.objects.using(DB_NAME).get(id=18)
        logo = "demo-logo.png"
    params = {"direction": 2}
    loc = None
    data = None
    if "location_id" in request.GET:
        try:
            loc = client.locations.get(id=request.GET["location_id"])
            data = loc.bordersaggregateddata_set.filter(
                    timeval__in=[8, 15, 24, 25, 26, 27])
            data = list(data.values("direction__order", "day",
                                    "phase", "timeval", "avg", "perc_85th"))
        except Location.DoesNotExist as e:
            pass
    locations = Client.objects.using(DB_NAME).get(id=15).locations.all()
    print("locations are", locations)

    if "borders" in request.get_full_path():
        template = loader.get_template(f'{BASE_URL}/borders-view.html')
    else:
        template = loader.get_template(f'{BASE_URL}/east-lothian-view.html')
    response = template.render({"location":  loc, "data": json.dumps(data), "logo": logo,
                                "locations": locations.as_geojson(), "client": client.name.lower().replace(" ", "-"),
                                "params":    params, "mapCode": "cjwgah7js2kbb1cp9aztbp6vm"}, request)
    # print(query_count_all())
    return HttpResponse(response)


def get_borders_data(request):
    if request.method == "POST":
        pass


@login_required(login_url=LOGIN_URL)
@require_POST
def get_ATC_overview(request):
    loc = Location.objects.using(DB_NAME).get(id=request.POST["locId"])
    data = loc.get_atc_overview()
    return JsonResponse({"status": "ok", "message": "Well, something happened anyway", "data": data})


@login_required(login_url=LOGIN_URL)
@require_POST
def get_ATC_scatter_plot(request):
    print("in scatter plot")
    loc = Location.objects.using(DB_NAME).get(id=request.POST["locId"])
    obs = Observation.objects.using(DB_NAME).filter(
            location=loc).prefetch_related("obsClass__obsClass", "direction")
    print(obs.query)
    data = ObservationSerializerForScatterChart(obs, many=True).data
    print("data is", data[:10])
    return JsonResponse({"status": "ok", "data": data})


@catch_database_errors
@login_required(login_url=LOGIN_URL)
def get_classed_volumes(request):
    client = request.user.client_set.all()[0]
    params = request.POST.dict()
    params["clientTz"] = client.timezone
    ids = json.loads(request.POST.get("ids", False))
    loc = Location.objects.using(DB_NAME).filter(id__in=ids)
    if 'project' in params:
        # params['project'] = loc[0].projects.all()[0].id
        project_loc = ProjectLocations.objects.using(DB_NAME).get(
                project_id=params['project'], location_id=loc[0])
        params['startDate'] = str(project_loc.startDate) + " 00:00"
        params['endDate'] = str(project_loc.endDate +
                                datetime.timedelta(days=1)) + " 00:00"
    obs_type = loc[0].observationType.id
    request.session.modified = True
    data = loc.get_classed_volumes(**params)
    return JsonResponse({"status": "ok", "data": data, 'obsType': obs_type})


@login_required(login_url=LOGIN_URL)
def get_weather_data(request):
    params = request.POST.dict()
    print("params are", params)
    loc = Location.objects.using(DB_NAME).get(id=params["id"])
    data = loc.get_weather_data(**params)
    return JsonResponse({"status": "ok", "data": data})


@login_required(login_url=LOGIN_URL)
def get_speed_data(request):
    loc = Location.objects.using(DB_NAME).get(id=request.POST["locId"])
    data = loc.get_speed_data()
    return JsonResponse({"status": "ok", "message": "Well, something happened anyway", "data": data})


@login_required(login_url=LOGIN_URL)
def toggle_bad_data(request):
    loc = Location.objects.using(DB_NAME).get(id=request.POST["location_id"])
    date = datetime.datetime.strptime(request.POST["date"], "%Y-%m-%d")
    hour = request.POST["hour"]
    Observation.objects.using(DB_NAME).toggle_bad_data(
            date=date, hour=hour, location=loc)
    return JsonResponse({"status": "ok"})


@login_required(login_url=LOGIN_URL)
def toggle_remove_day(request):
    loc = Location.objects.using(DB_NAME).get(id=request.POST["location_id"])
    date = datetime.datetime.strptime(request.POST["date"], "%Y-%m-%d")
    Observation.objects.using(DB_NAME).toggle_remove_day(
            date=date, location=loc)
    return JsonResponse({"status": "ok"})


#######################################################################################################
#
#
# Dashboard functions
#
#
#######################################################################################################
@login_required(login_url=LOGIN_URL)
def get_crt_style_data(request):
    print("params in crt data  are", request.POST)
    kwargs = request.POST.dict()
    try:
        kwargs["weekday"] = int(kwargs["weekday"])
    except Exception as e:
        pass
    if "table" in kwargs:
        # link a temp table to the clients session
        kwargs["table"] = request.session.session_key
    locations = json.loads(request.POST["ids"])
    locations = Location.objects.using(DB_NAME).filter(id__in=locations)
    print("loaded locations in view")
    table, headers = locations.as_crt_style_table()
    return JsonResponse({"status":    "ok", "data": locations.get_classed_volumes(**kwargs),
                         "table":     table, "headers": headers,
                         "classes":   locations[0].classes.all().as_dashboard_class_list(),
                         "dateRange": locations.get_date_range()})


def camden_speed_view(request):
    template = loader.get_template(f'{BASE_URL}/camden-speeds.html')
    classes = [43087, 43088, 43089, 43090, 43091, 43092, 43093, 43076]
    client = request.user.client_set.all(
    ).prefetch_related("locations__area")[0]
    kwargs = request.GET.dict()
    id = kwargs.pop("id", None)
    year = kwargs.pop("year", datetime.datetime.now().year)
    # print(query_count_all())
    if id:
        loc = Location.objects.using(DB_NAME).get(id=id)
        dates = AssociatedObservation.objects.using(DB_NAME).filter(location_id=id, date__year=year,
                                                                    obsClass_id__in=classes)
        dates = dates.annotate(d=TruncDate("date")).values_list(
                "d", flat=True).distinct()
        dates = [d.strftime("%Y-%m-%d") for d in dates]
    else:
        dates = []
        loc = None
    # print(query_count_all())
    response = template.render({"calendarDates": dates, "client": client, "id": id, "loc": loc,
                                "view":          ClientView.objects.using(DB_NAME).get(id=39),
                                "year":          year}, request)

    return HttpResponse(response)


@login_required(login_url=LOGIN_URL)
def get_camden_speed_data(request):
    print("params in camden speed data  are", request.POST)
    rng = pd.date_range('2011-03-01 00:00', '2011-03-01 23:55',
                        freq="5T", ambiguous=True)
    kwargs = request.POST.dict()
    id = kwargs.pop("id", None)
    try:
        loc = Location.objects.using(DB_NAME).get(id=id)
    except Exception as e:
        loc = Location.objects.using(DB_NAME).get(id=16207)
    classes = [43087, 43088, 43089, 43090, 43091, 43092, 43093, 43076]
    d = datetime.datetime.strptime(kwargs["date"], "%Y-%m-%d")
    timezone.activate("Europe/London")
    print("aware date is now", timezone.make_aware(
            d, timezone=pytz.timezone("Europe/London")))
    d1 = timezone.localtime(timezone.make_aware(d))
    current_tz = timezone.get_current_timezone()
    data = AssociatedObservation.objects.using(DB_NAME).filter(location=loc, date__date=kwargs["date"],
                                                               obsClass_id__in=classes).values_list(
            "date__time", "location_id", "obsClass_id", "direction_id", "value")
    df = pd.DataFrame(
            data, columns=["time", "location", "classorder", "directionorder", "value"])
    df.set_index(["time", "location", "classorder",
                  "directionorder"], inplace=True)
    index = pd.MultiIndex.from_product([[r for r in rng.time], [loc.id], classes, [4, 3]],
                                       names=["time", "location", "classorder", "directionorder"])
    df = df.reindex(index, fill_value="-").reset_index()
    df.columns = ["time", "location", "classorder", "directionorder", "value"]
    df["classorder"] = df["classorder"].apply(lambda x: classes.index(x))
    data = df.groupby(["classorder", "directionorder"]).agg({"value": list})
    data = [data.iloc[data.index.get_level_values('directionorder') == 4]["value"].tolist(),
            data.iloc[data.index.get_level_values('directionorder') == 3]["value"].tolist()]
    return JsonResponse({"status": "ok", "data": data, "graphLabels": [r[:5] for r in rng.time.astype(str)],
                         "title":  loc.area.name + " " + loc.name + " for " + datetime.datetime.strptime(kwargs["date"],
                                                                                                         "%Y-%m-%d").strftime(
                                 "%a %d/%m/%Y")})


@login_required(login_url=LOGIN_URL)
def start_factoring_thread(request):
    timezone.activate("UTC")
    numLocs = Location.objects.using(DB_NAME).filter(
            temp=1, factoringEdited=1).count()
    if numLocs == 0:
        return HttpResponseBadRequest(reason="No changes made since last factoring, process exited")
    try:
        event = FactoringEvent.objects.create(
                startDate=timezone.make_aware(
                        datetime.datetime.utcnow().replace(microsecond=0)),
                numLocations=numLocs,
                eventType="User manually started factoring process")
    except django.db.IntegrityError as e:
        return HttpResponseBadRequest(
                reason="Cannot apply factoring at this time, another factoring process is currently active")
    threading.Thread(target=factoring_thread, args=(event,)).start()
    return JsonResponse({"status": "ok"})


def factoring_thread(event):
    timezone.activate("UTC")
    locs = Location.objects.using(DB_NAME).filter(temp=1, factoringEdited=1)
    try:
        for loc in locs:
            startDate = Observation.objects.using(DB_NAME).filter(
                    location=loc, status=0).latest("date").date
            startOfNextMonth = timezone.make_aware(datetime.datetime(startDate.year + (startDate.month // 12),
                                                                     ((startDate.month % 12) + 1), 1))
            endDate = timezone.make_aware(
                    datetime.datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0))
            params = {"applyStart": startOfNextMonth,
                      "applyEnd":   endDate}
            loc.processingFactoring = True
            loc.save()

            loc.apply_factoring(**params)
            loc.processingFactoring = False
            loc.factoringEdited = False
            loc.save()
        msg = "Successfully completed manual factoring process"
    except Exception as e:
        msg = "Manual factoring process failed on location " + \
              str(loc.id) + " " + traceback.format_exc()
        sendMail(SEND_MAIL_TO, "Factoring event failure", msg)
    finally:
        event.endDate = timezone.make_aware(
                datetime.datetime.utcnow().replace(microsecond=0))
        event.msg = msg
        event.save()


def get_factoring_data(request):
    print("params are", request.POST)
    kwargs = request.POST.dict()
    kwargs["table"] = request.session.session_key + \
                      "_factoring_"  # link a temp table to the clients session
    locations = json.loads(request.POST["ids"])
    locations = Location.objects.using(DB_NAME).filter(id__in=locations)
    table, headers = locations.as_crt_style_table()
    return JsonResponse({"status":  "ok", "data": locations.get_classed_volumes(**kwargs),
                         "table":   table, "headers": headers,
                         "classes": locations[0].ordered_class_queryset().as_dashboard_class_list()})


@login_required(login_url=LOGIN_URL)
def redirect_to_view(request):
    params = request.GET.copy()
    print("in redirect, params are", params)
    request.session["params"] = params
    return HttpResponseRedirect(f"/{BASE_URL}/dashboard")


def logon_view(request, client):
    # print("in logon view, next is", request)
    # print("in logon view, next is", request.GET.get('next'))
    try:
        print('logon_view')
        print("client is -", client)
        # client = 'eastlothian'
        client = Client.objects.using(DB_NAME).get(nameForUrl=client)
        print("client2 is -", client)
    except Client.DoesNotExist as e:
        print("error", e)
        client = Client(
                logonBackground=f"{BASE_URL}/roadTraffic.jpg", nameForUrl="")
    template = loader.get_template(f'{BASE_URL}/logonScreen.html')
    response = template.render({"client": client}, request)
    # print('response', response)
    # return HttpResponse(response)
    print("client test - ", client)
    return render(request, f'{BASE_URL}/logonScreen.html', {"client": client})


def admin_logon_view(request):
    client = Client(
            logonBackground=f"{BASE_URL}/LandingPage.png", nameForUrl="")
    template = loader.get_template(f'{BASE_URL}/adminLogonScreen.html')
    response = template.render({"client": client}, request)
    return HttpResponse(response)


def logon(request, client):
    print("next is", request.GET.get('next'))
    data = request.POST
    if "username" in data and "password" in data:
        username = data["username"]
        pw = data["password"]
        user = authenticate(request, username=username, password=pw)
        print("user is", user)
        if user is None:
            print("authentication failed")
            return JsonResponse({"status": "error", "message": "Incorrect User Email or Password!"})
            # return HttpResponseRedirect(f"/{BASE_URL}" + client)
        login(request, user, backend='aecon.myauthbackend.CustomAuthBackend')
    else:
        if hasattr(request, "user"):
            # print("already logged in",request.user)
            user = request.user
        else:
            return HttpResponseForbidden()
    if "admin" in data:
        print(request.user, "trying to log on to admin")
        if request.user.groups.filter(name="admin").exists():
            redirect_string = f"/{BASE_URL}/admin/allLocations"
        else:
            return HttpResponseForbidden()
    else:
        try:
            client = user.client_set.all()[0]
            pass
        except Exception as e:
            return HttpResponseRedirect(f"/{BASE_URL}" + client)
        redirect_string = f"/{BASE_URL}/map-full-view"
    if request.GET.get("next"):
        redirect_string = request.GET.get("next")
        try:
            resolve_match = django.urls.resolve(redirect_string)
        except django.urls.Resolver404:
            redirect_string = f"/{BASE_URL}/welcome-page"
    # do something on bad input
    print("after logon, redirecting to", redirect_string)
    return JsonResponse({"status": "success", "redirect_string": redirect_string})
    # return redirect(redirect_string)


@login_required(login_url=LOGIN_URL)
def user_logout(request):
    print("logging out", request.user, type(request.build_absolute_uri()))
    client = request.user.client_set.all()[0]
    if request.user.groups.filter(name="admin").exists() and "admin" in request.build_absolute_uri():
        redirect_string = f"/{BASE_URL}/admin"
    else:
        redirect_string = f"/{BASE_URL}/"
    logout(request)
    return redirect(redirect_string)


@catch_database_errors
@login_required(login_url=LOGIN_URL)
def dashboard(request):
    if "params" in request.session:
        print("in dashboard, params are", request.session["params"])
        try:
            client = request.user.client_set.all(
            ).prefetch_related("locations__area")[0]
            if "params" not in request.session:
                return redirect(f"/{BASE_URL}/trafic-control")
            try:
                view = ClientView.objects.using(DB_NAME).select_related("view").get(
                        id=request.session["params"]["view"])
            except:
                view = ClientView.objects.using(
                        DB_NAME).select_related("view").get(id='53')
            if view.view.redirect:
                return redirect(f"/{BASE_URL}/" + view.view.html_file_name)

            template = loader.get_template(
                    f'{BASE_URL}/' + str(view.view.html_file_name) + '.html')
            # , "location": loc, "DEBUG": settings.DEBUG}
            context = {"view": view, "client": client}
            if "location_id" in request.session["params"]:
                print("in dashboard location id is",
                      request.session["params"]["location_id"])
                loc = Location.objects.using(DB_NAME).get(
                        id=request.session["params"]["location_id"])
                context["location"] = loc
            else:
                loc = None
            response = template.render(context, request)
            # print(query_count_all())
            return HttpResponse(response)
        except Client.DoesNotExist as e:
            logout(request)
            return redirect(f"/{BASE_URL}")
        except Location.DoesNotExist as e:
            return redirect(f"/{BASE_URL}/trafic-control")
    return redirect(f"/{BASE_URL}")


@login_required(login_url=LOGIN_URL)
def view_project(request):
    params = request.POST
    # print("params are",params)
    if "project" not in params:
        return HttpResponseBadRequest(reason="Invalid Request")
    try:
        proj = Project.objects.using(DB_NAME).get(id=params["project"])
        locs = proj.get_project_locations_by_obstype()
        request.session["project"] = proj.id
        template = loader.render_to_string(
                f'{BASE_URL}/initial-user-view.html', {"locations": locs, "project": proj})
        return JsonResponse({"data": template})
    except Project.DoesNotExist as e:
        return HttpResponseBadRequest(reason="No such project")
    except django.db.utils.Error as e:
        sendMail(SEND_MAIL_TO, "School_Streets Server Error",
                 "view_project , django error " + str(e))
        return HttpResponseServerError
    except pymysql.Error as e:
        sendMail(SEND_MAIL_TO, "School_Streets Server Error",
                 "view_project , mysql error " + str(e))
        return HttpResponseServerError


@login_required(login_url=LOGIN_URL)
def view_location(request):
    if "id" not in request.GET:
        return HttpResponseBadRequest("No Location Specified")
    try:
        loc = Location.objects.using(DB_NAME).get(id=request.GET["id"])
        request.session["location_id"] = loc.id
        proj = loc.project_set.all()[0]
        projects = Project.objects.using(DB_NAME).all()
        if loc.observationType.id == 4:
            template = loader.get_template(f'{BASE_URL}/atc-view.html')
        sideBar = Location.objects.using(DB_NAME).filter(
                project=proj).format_for_sidebar(request)
        response = template.render(
                {"sideBar": sideBar, "proj": proj, "loc": loc, "projects": projects}, request)
        return HttpResponse(response)
    except Location.DoesNotExist as e:
        return HttpResponseBadRequest(reason="No such location")
    except django.db.utils.Error as e:
        sendMail(SEND_MAIL_TO, "School_Streets  Server Error",
                 "view_location , django error " + str(e))
        return HttpResponseServerError
    except pymysql.Error as e:
        sendMail(SEND_MAIL_TO, "School_Streets Server Error",
                 "view_location , mysql error " + str(e))
        return HttpResponseServerError


@catch_database_errors
# @login_required(login_url=LOGIN_URL)
def get_client_locations(request):
    print("in get client locations")
    # proj_loc = ProjectLocations.objects.using(DB_NAME).filter(location=loc[0]).order_by('startDate')
    # project_lst = []
    # for item in proj_loc:
    #     project = Project.objects.using(DB_NAME).get(id=item.project_id)
    #     project_lst.append({'id': project.id, 'name': project.name, "start_date": item.startDate)

    if request.user.is_authenticated:
        client = request.user.client_set.all()[0]
        locs = client.locations.all().exclude(virtual=1)
        print(f"Authenticated client locations: {locs.count()}")
    else:
        client = Client.objects.using(DB_NAME).get(id=18)
        locs = client.locations.filter(observationType_id=9).exclude(virtual=1)
        print(f"Client ID 18 locations: {locs.count()}")

    if "exclude associated" in request.POST and request.POST["exclude associated"].lower() == "true":
        locs = locs.exclude_associated_locations()
    # print("here")
    data = {}
    for loc in locs:
        proj_loc = ProjectLocations.objects.using(
                DB_NAME).filter(location=loc).order_by('startDate')
        projects = []

        if loc.observationType.id == 9 or loc.observationType.id == 1 or loc.observationType.id == 10 or loc.observationType.id == 12 or loc.observationType.id == 11:
            for item in proj_loc:
                project = Project.objects.using(
                        DB_NAME).get(id=item.project_id)
                projects.append(
                        {'id': project.id, 'name': project.project_no, "start_date": item.startDate})

        data[loc.id] = {
                'obs_type_id': loc.observationType.id, "projects": projects}
    print(locs[0])
    print(locs[0].classes.all())
    response = {"status":  "ok", "locations": locs.as_geojson(),
                "classes": locs[0].classes.all().as_dashboard_class_list(),
                "data":    data
                }
    for key in response:
        print(f'{key}: {response[key]}')

    return JsonResponse({"status":  "ok", "locations": locs.as_geojson(),
                         "classes": locs[0].classes.all().as_dashboard_class_list(),
                         "data":    data
                         })


@login_required(login_url=LOGIN_URL)
def get_tracks(request):
    params = request.POST
    # print("params are", params)
    if "location_id" not in params:
        return HttpResponseBadRequest("No Location Specified")
    try:
        loc = Location.objects.using(DB_NAME).get(id=params["location_id"])
        tracks = loc.get_tracks("", "")
        return JsonResponse({"tracks": tracks})
    except Location.DoesNotExist as e:
        print("no such location", params["location_id"])
        return HttpResponseBadRequest(reason="No such location")
    except django.db.utils.Error as e:
        sendMail(SEND_MAIL_TO, "School_Streets Server Error",
                 "view_location , django error " + str(e))
        return HttpResponseServerError
    except pymysql.Error as e:
        sendMail(SEND_MAIL_TO, "School_Streets Server Error",
                 "view_location , mysql error " + str(e))
        return HttpResponseServerError


@catch_database_errors
@login_required(login_url=LOGIN_URL)
def events(request):
    if request.method == "GET":
        data = request.GET.dict()
        eventId = None
    elif request.method == "POST":
        data = request.POST.dict()
        print("data in save event is", data)
        try:
            event = Event.objects.create(**data)
            event.addedBy = request.user
            event.save(using=DB_NAME)
            eventId = event.id
        except ValidationError as e:
            print(e)
            return HttpResponseBadRequest(reason=e.messages[0])
        except KeyError as e:
            return HttpResponseBadRequest(reason="missing field - " + str(e))
        except Exception as e:
            return HttpResponseServerError(reason=str(e))
    events = Event.objects.using(DB_NAME).all()
    if "startDate" in data:
        events = events.filter(
                date__gte=data["startDate"], date__lt=data["endDate"])
    if "location_id" in data:
        events = events.filter(location_id=data["location_id"])
    return JsonResponse({"events": events.order_by("-date").as_html(), "event_id": eventId})


@catch_database_errors
@login_required(login_url=LOGIN_URL)
def get_events(request):
    data = request.POST.dict()


@catch_database_errors
@login_required(login_url=LOGIN_URL)
def save_clustering(request):
    if FactoringEvent.objects.get_current_event() is not None:
        return HttpResponseBadRequest(reason="There is a factoring process currently running. You cannot make changes"
                                             " while the process is active, changes were not saved")
    locations = json.loads(request.POST["ids"])
    clustering = json.loads(request.POST["clustering"])
    print("clustering is", clustering)
    Location.objects.using(DB_NAME).filter(
            id__in=locations).set_clusters(clustering)
    return JsonResponse({"status": "hi"})


@catch_database_errors
@login_required(login_url=LOGIN_URL)
def get_clustering(request):
    locations = json.loads(request.POST["ids"])
    result = Location.objects.using(DB_NAME).filter(
            id__in=locations).get_clusters()
    return JsonResponse({"data": result})


def get_perm_clustering_for_temp_site(request):
    # print("params are",request.POST)
    id = request.POST["id"]
    client = request.user.client_set.all()[0]
    return JsonResponse({"data": client.clustering_as_table(id)})


#######################################################################################################
#
#
# deal with data downloads
#
#
#######################################################################################################

def create_averages_df(locs, **kwargs):
    if kwargs["period"] == "D":
        kwargs["period"] = 1440
    else:
        kwargs["period"] = int(kwargs["period"])
    kwargs["weekday"] = 4
    print("from created Average df")
    df = locs.get_data_from_crt_averages_table(**kwargs)
    df = df.replace("-", -1)
    df.loc[df["day"] == 0, "day"] = "Weekday"
    df.loc[df["day"] == 1, "day"] = "Saturday"
    df.loc[df["day"] == 2, "day"] = "Sunday"

    def convert_time(row):
        d = datetime.timedelta(minutes=kwargs["period"] * row["seg"])
        secs = d.total_seconds()
        hours = int(secs / 3600)
        minutes = int(secs / 60) % 60
        return datetime.datetime(2020, 1, 1, hours, minutes).time()

    df["Date"] = df.apply(convert_time, axis=1)
    df = pd.pivot_table(df, values="value", columns="classorder",
                        index=["location", "day", "Date", "directionorder"]).reset_index()
    df = df.replace("-", np.nan)
    df = df.replace(-1, np.nan)

    df = locs.format_dataframe_for_download(df, **kwargs)
    return df


def create_hourly_df(locs, **kwargs):
    df = locs.download_totals(**kwargs)
    print("df is", df)
    df = df.replace("-", np.nan)
    df = df.groupby(["Area", "Site", 'Time', "Direction"],
                    as_index=False).mean()
    df = df.replace(np.nan, "-")
    print("and now df is", df)
    return df


def create_regional_report(locs, file, **kwargs):
    # print("kwargs are",kwargs)
    df = create_averages_df(locs, **kwargs)
    df["extra"] = df.iloc[:, 6:-1].sum(axis=1)
    dailyDf = create_hourly_df(locs, **kwargs)
    dailyDf = dailyDf.replace("-", 0)
    dailyDf["extra"] = dailyDf.iloc[:, 4:-1].sum(axis=1)
    row = 12
    wb = openpyxl.load_workbook(
            os.path.join(settings.BASE_DIR, f"{BASE_URL}/NEW Blank Regional Reporting Template.xlsm"), keep_vba=True)
    wb.active = 0
    sheet = wb.active
    sheet.sheet_view.topLeftCell = "A1"
    sheet["C24"].value = kwargs["region"]
    sheet["J29"].value = kwargs["startDate"].strftime("%B %Y")
    currentYearStart = kwargs["startDate"].replace(day=1, month=1)
    previousYearStart = currentYearStart.replace(
            year=currentYearStart.year - 1)
    for loc in locs:
        wb.active = 1
        sheet = wb.active
        sheet.sheet_view.topLeftCell = "A1"
        sheet.cell(row=row, column=3).value = loc.name
        temp = df[df["Site"] == loc.id]
        sheet.cell(row=row, column=9).value = int(
                temp.loc[temp["day"] == "Weekday"]["Pedestrian"])
        sheet.cell(row=row, column=10).value = int(
                temp.loc[temp["day"] == "Weekday"]["Cyclist"])
        sheet.cell(row=row, column=11).value = int(
                temp.loc[temp["day"] == "Weekday"]["extra"])
        sheet.cell(row=row, column=14).value = int(
                temp.loc[temp["day"] == "Saturday"]["Pedestrian"])
        sheet.cell(row=row, column=15).value = int(
                temp.loc[temp["day"] == "Saturday"]["Cyclist"])
        sheet.cell(row=row, column=16).value = int(
                temp.loc[temp["day"] == "Saturday"]["extra"])
        sheet.cell(row=row, column=19).value = int(
                temp.loc[temp["day"] == "Sunday"]["Pedestrian"])
        sheet.cell(row=row, column=20).value = int(
                temp.loc[temp["day"] == "Sunday"]["Cyclist"])
        sheet.cell(row=row, column=21).value = int(
                temp.loc[temp["day"] == "Sunday"]["extra"])

        temp = dailyDf[dailyDf["Site"] == loc.id]
        sheet.cell(row=row, column=24).value = int(temp.iloc[0]["Pedestrian"])
        sheet.cell(row=row, column=25).value = int(temp.iloc[0]["Cyclist"])
        sheet.cell(row=row, column=26).value = int(temp.iloc[0]["extra"])

        wb.active = 2
        sheet = wb.active
        sheet.sheet_view.topLeftCell = "A1"
        sheet.cell(row=row, column=3).value = loc.name
        sheet["I8"].value = previousYearStart.year
        sheet["M8"].value = previousYearStart.year
        sheet["Q8"].value = previousYearStart.year

        sheet["J8"].value = currentYearStart.year
        sheet["N8"].value = currentYearStart.year
        sheet["R8"].value = currentYearStart.year

        ##
        # daily average
        ##
        print("filling previous year daily average",
              previousYearStart, currentYearStart)

        val = loc.dailytotals_set.filter(date__year=previousYearStart.year)
        val = val.values("location").annotate(
                total=Avg("value")).values_list("total", flat=True)
        print("val for ", loc.id, val)
        sheet.cell(row=row, column=9).value = round(
                val[0]) if len(val) > 0 else "-"

        print("filling current year daily average",
              previousYearStart, currentYearStart)
        val = loc.dailytotals_set.filter(date__year=currentYearStart.year)
        val = val.values("location").annotate(
                total=Avg("value")).values_list("total", flat=True)
        print("val for ", loc.id, val)
        sheet.cell(row=row, column=10).value = round(
                val[0]) if len(val) > 0 else "-"

        ##
        # previous year to date
        ##
        print("filling previous year to date", previousYearStart,
              kwargs["endDate"].replace(year=previousYearStart.year))
        val = loc.dailytotals_set.filter(date__gte=previousYearStart,
                                         date__lt=kwargs["endDate"].replace(year=previousYearStart.year))
        print("val is", val)
        val = val.values("location").annotate(
                total=Sum("value")).values_list("total", flat=True)
        print("val for ", loc.id, val)
        sheet.cell(row=row, column=13).value = val[0] if len(val) > 0 else "-"
        ##
        # current year to date
        ##
        print("filling current year to date",
              currentYearStart, kwargs["endDate"])
        val = loc.dailytotals_set.filter(
                date__gte=currentYearStart, date__lt=kwargs["endDate"])
        val = val.values("location").annotate(
                total=Sum("value")).values_list("total", flat=True)
        print("val for ", loc.id, val)
        sheet.cell(row=row, column=14).value = val[0] if len(val) > 0 else "-"
        ##
        # previous year moving total
        ##
        print("filling prev year moving", kwargs["startDate"].replace(year=previousYearStart.year - 1),
              kwargs["endDate"].replace(year=previousYearStart.year))
        val = loc.dailytotals_set.filter(date__gte=kwargs["startDate"].replace(year=previousYearStart.year - 1),
                                         date__lt=kwargs["endDate"].replace(year=previousYearStart.year))
        val = val.values("location").annotate(
                total=Sum("value")).values_list("total", flat=True)
        print("val for ", loc.id, val)
        sheet.cell(row=row, column=17).value = val[0] if len(val) > 0 else "-"
        ##
        # current year moving total
        ##
        print("filling current year moving", kwargs["startDate"].replace(year=previousYearStart.year),
              kwargs["startDate"])
        val = loc.dailytotals_set.filter(date__gte=kwargs["startDate"].replace(year=previousYearStart.year),
                                         date__lt=kwargs["endDate"])
        val = val.values("location").annotate(
                total=Sum("value")).values_list("total", flat=True)
        print("val for ", loc.id, val)
        sheet.cell(row=row, column=18).value = val[0] if len(val) > 0 else "-"

        row += 1
    wb.active = 0
    sheet = wb.active
    print("saving file to", file)
    wb.save(os.path.join(settings.BASE_DIR, "data", file))


def download_thread(file, **kwargs):
    token = file
    errorFile = os.path.join(settings.BASE_DIR, "data",
                             file.split(".")[0] + "_failed.txt")
    msg = None
    emailDets = None
    # print("kwargs are",kwargs)
    try:
        locs = Location.objects.using(DB_NAME).filter(id__in=kwargs["ids"])
        if not os.path.exists(os.path.join(settings.BASE_DIR, "data")):
            os.mkdir(os.path.join(settings.BASE_DIR, "data"))
        if kwargs["calc"] == "regional":
            file = file + ".xlsm"
            areas = Area.objects.using(DB_NAME).filter(region=kwargs["region"])
            locs = kwargs["client"].locations.filter(
                    temp=0, area__in=areas, observationType_id=1)
            locs = VivacityAPI.objects.using(DB_NAME).get(
                    id=2).locations.filter(id__in=locs)
            kwargs["direction"] = "combined"
            kwargs["period"] = "D"
            kwargs["classes"] = list(
                    locs[0].classes.values_list("obsClass_id", flat=True))
            kwargs["startDate"] = kwargs["startDate"].replace(day=1)
            kwargs["endDate"] = (kwargs["startDate"] +
                                 datetime.timedelta(days=35)).replace(day=1)
            create_regional_report(locs, file, **kwargs)
        elif kwargs["calc"] == "weather":
            file = file + ".csv"
            df = locs.download_weather(**kwargs)
            df.to_csv(os.path.join(settings.BASE_DIR, "data", file))
        else:

            file = file + ".csv"
            dataframes = []
            for loc in locs:
                print("processing", loc)
                selectedLocs = Location.objects.using(
                        DB_NAME).filter(id=loc.id)
                print("selected locs are", selectedLocs)
                if kwargs["calc"] == "averages":
                    df = create_averages_df(selectedLocs, **kwargs)
                if kwargs["calc"] == "hourly_averages":
                    df = create_hourly_df(selectedLocs, **kwargs)
                if kwargs["calc"] == "totals":
                    df = selectedLocs.download_totals(**kwargs)
                df = df.loc[:, ~df.columns.duplicated()]
                dataframes.append(df)
            df = pd.concat(dataframes)
            # print(df.info())
            locsDict = {s.id: s for s in locs}
            print("now df is", df, lambda x: locsDict[x].name)
            df["Site"] = df["Site"].apply(lambda x: locsDict[x].name)
            df.to_csv(os.path.join(settings.BASE_DIR, "data", file))

    except PermissionError as e:
        msg = "Error creating file " + str(type(e)) + str(e)
    # except KeyError as e:
    #   msg = "Error Writing to Template " + str(type(e)) + str(e)
    except pymysql.Error as e:
        msg = "Database Error " + str(type(e))
    except django.db.utils.Error as e:
        msg = "Django Error " + str(type(e)) + str(e)
    # except Exception as e:
    #    msg = "Generic Error " + str(type(e)) + str(e)
    print("msg is", msg, errorFile)
    if msg:
        emailDets = (
                "Tracsis Data Download Failed",
                "Your data download failed to complete. This was due to the following error \n"
                + msg + "\n.Please contact Tracsis with this information")
        sendMail(SEND_MAIL_TO, *emailDets)
        with open(errorFile, "w") as f:
            f.write(msg)
    if "email" in kwargs and kwargs["email"] != "":
        survey_type = str(kwargs["dataType"]).capitalize()
        root_url = kwargs["rootUrl"]

        url = f'{root_url}downloadFile?token={token}&name=Tracsis%20{survey_type}%20{str(datetime.datetime.now().date())}&extension=csv'
        email = kwargs["email"]
        content = "<p>Dear " + kwargs["username"] + ",</p>"
        content += f"<p>The {survey_type} Data you requested from AECON is ready. Please note that, <a href='{url}'> link to the data, this link is available for 15 days only.</a></p>"
        content += "<p>Please don’t hesitate to contact us, If you have any issue.</p>"
        content += "<p>Kind Regards,"
        content += "<br>AECON Team</p>"
        content += "<p>ttd.admin@tracsis.com</p>"
        sendMail(email, 'Data Download is Ready', content)
        print('Email Sent')
        # sendMail(kwargs["email"], *emailDets, sender=None)


@catch_database_errors
@login_required(login_url=LOGIN_URL)
def start_download_process(request):
    params = request.POST.dict()
    client = request.user.client_set.all()[0]
    try:
        params["startDate"] = datetime.datetime.strptime(
                params["startDate"], "%Y-%m-%d")
        params["endDate"] = datetime.datetime.strptime(
                params["endDate"], "%Y-%m-%d")
        params["classes"] = [int(c) for c in json.loads(params["classes"])]
        params["ids"] = json.loads(params["ids"])
        print("params are", type(params['ids']))
        params["client"] = client

        if client.id != 12:
            params["report_type"] = "Countlines"

        datapoints = calculate_num_datapoints(**params)
        if datapoints > 5000000:
            return HttpResponseBadRequest(
                    reason="You have requested too much data in a single request. Please break your request down into smaller downloads")
        if Location.objects.using(DB_NAME).filter(id__in=params["ids"]).values(
                "observationType").distinct().count() > 1:
            return HttpResponseBadRequest(reason="You cannot download data for multiple different sensor types")
        assert client.locations.filter(
                id__in=params["ids"]).count() == len(params["ids"])
        file = generate_salt(16) + str(datetime.datetime.now().date())
        params["clientTz"] = client.timezone
        params["table"] = request.session.session_key + "_download"
        params["username"] = request.user.first_name
        threading.Thread(target=download_thread,
                         args=(file,), kwargs=params).start()
        msg = "Your download will start shortly"
        status = "download"
        if "email" in params and params["email"] != "":
            msg = "A link to your download will be sent to " + \
                  str(params["email"])
            status = "email"
        return JsonResponse({"status": status, "filename": file, "message": msg})
    except ValueError:
        return HttpResponseBadRequest(reason="Invalid or missing date")
    except KeyError as e:
        return HttpResponseBadRequest(reason="Missing field : " + str(e))
    except AssertionError:
        return HttpResponseBadRequest(reason="You do not have access to the data for one or more selected sites")


def download_file(request):
    data = request.GET.dict()
    token = data['token']
    name = data.get('name', 'null')
    extension = data['extension']
    if extension in ['xlsx', 'xls']:
        content_type = 'application/vnd.ms-excel'
    elif extension == 'zip':
        content_type = 'application/zip'
    elif extension == 'csv':
        content_type = 'text/csv'
    else:
        return HttpResponseBadRequest('unknown file extension')

    if extension != 'csv':
        path = os.path.join(EXCEL_ROOT, token)
    else:
        path = os.path.join(settings.BASE_DIR, "data", token + '.csv')

    try:
        with open(path, 'rb') as fh:
            response = HttpResponse(fh, content_type=content_type)
            response['Content-Disposition'] = 'inline;filename=' + \
                                              name + '.' + extension
            return response
    except Exception as e:
        print(e)
        return HttpResponseBadRequest('unknown file')


@login_required(login_url=LOGIN_URL)
def check_file_ready(request):
    fileName = request.POST["file"]
    params = request.POST.dict()
    if 'upload_status' not in params:
        tail = fileName.split(".")[0]
        print("fileName is", fileName, "tail is", tail)
        # fileName = os.path.join(settings.BASE_DIR,"data", fileName)
        for ext in [".csv", ".xlsm"]:
            if os.path.exists(os.path.join(settings.BASE_DIR, "data", fileName + ext)):
                return JsonResponse({"status": "ready", "fileToken": fileName})
        if os.path.exists(os.path.join(settings.BASE_DIR, "data", tail + "_failed.txt")):
            with open(os.path.join(settings.BASE_DIR, "data", tail + "_failed.txt"), "r") as f:
                msg = f.readline()
            return JsonResponse({"status": "failed", "message": msg})
        return JsonResponse({"status": "not ready"})
    else:
        response = {'status': 'started', 'message': 'file not ready'}

        if os.path.exists(os.path.join(EXCEL_ROOT, f'{fileName}_failed.txt')):
            with open(os.path.join(EXCEL_ROOT, f'{fileName}_failed.txt'), 'r') as f:
                msg = f.readline()
            response = {"status": "failed", "message": msg}
        elif os.path.exists(os.path.join(EXCEL_ROOT, f'{fileName}.txt')):
            with open(os.path.join(EXCEL_ROOT, f'{fileName}.txt'), 'r') as f:
                msg = f.readlines()
                response['status'] = 'processing'
                if len(msg) and 'Uploading Process Completed' in msg[-1]:
                    response['status'] = 'success'
                response['message'] = json.dumps(msg)
        return JsonResponse(response)


def reset_password(request):
    if request.method == "POST":
        print(request.POST)
        if "Email" in request.POST:
            email = request.POST["Email"]
            try:
                user = User.objects.using(DB_NAME).get(email=email)
            except User.DoesNotExist as e:
                return JsonResponse({"status": "error", "user_id": None, "message": "Email not Registered"})
            otp = random.randint(100001, 999999)

            try:
                date_time = timezone.make_aware(
                        datetime.datetime.utcnow().replace(microsecond=0))
                # save_otp = Otp(user_id = user, otp=otp,created_date_time = datetime.datetime.now().replace(microsecond=0))
                save_otp = Otp(user_id=user, otp=otp,
                               created_date_time=date_time)
                save_otp.save(using=DB_NAME)
            except Exception as e:
                print(e)
                return JsonResponse({"status": "error", "message": "Problem in Creating OTP"})
            try:
                print(user)
                link = "http://" + request.get_host() + "/aecon/reset_password?id=" + str(otp)
                content = "<p>Dear " + str(user.get_short_name()) + ",</p>"
                content += "<p>You have received this email because you have been added as a new user of the AECON Traffic Speed and Volume Dashboard or you are already a user and have requested to reset the password for your user account.</p>"
                content += "<p> Account username: " + str(user.email) + "."
                content += "<br>To reset your password, please click the link below. </p>"
                content += "<p> <a href ='" + link + "'>Click here to reset password</a></p>"
                content += f"<br><p>Please note this link will expire after 48 hours. Once expired you can request a new link to reset your password via the website. <a href ='http://{request.get_host()}/'>http://{request.get_host()}/</a></p>"

                content += "<br><p>Regards,"
                content += "<br>AECON Team</p>"

                sendMail(email, "Password Reset", content)

            except Exception as e:
                print(e)
                return JsonResponse({"status": "error", "message": "Problem in sending mail"})
            return JsonResponse(
                    {"status":  "success", "user_id": user.id,
                     "message": "Please check your email for reset password link"})
        elif "OTP" in request.POST:
            otp = request.POST["OTP"]
            user_id = request.POST["user_id"] if request.POST["user_id"] else None
            otp = Otp.objects.using(DB_NAME).filter(
                    user_id=user_id, otp=otp).order_by("created_date_time").last()
            if otp is None:
                return JsonResponse({"status": "error", "message": "Please Enter Valid Otp "})
            else:
                current_time = timezone.make_aware(
                        datetime.datetime.utcnow().replace(microsecond=0))
                time_diff = current_time - otp.created_date_time
                time_diff = time_diff.total_seconds()
                time_diff = round(time_diff / 60)
                if time_diff <= 5:
                    return JsonResponse({"status": "success", "message": "Otp Verified Successfully", "id": otp.id})
                else:
                    return JsonResponse({"status": "error", "message": "Otp Expired"})

            # return JsonResponse({"status": "sucess", "message": "Otp Verified Successfully"})
        elif "Password" in request.POST and "ConfirmPassword" in request.POST:
            password = request.POST["Password"]
            cnf_password = request.POST["ConfirmPassword"]
            try:
                user = User.objects.using(DB_NAME).get(
                        id=request.POST["user_id"])
            except User.DoesNotExist as e:
                return JsonResponse({"status": "erorr", "message": "User Dosen't Exist"})
            if password == cnf_password:
                try:
                    user.set_password(password)
                    user.save(using=DB_NAME)
                    return JsonResponse({"status": "success", "message": "Password Changes Successfully"})
                except Exception as e:
                    return JsonResponse({"status": "error", "message": "Error in changing Password"})
            else:
                return JsonResponse({"status": "error", "message": "Password Doesnot Match"})
        else:
            return JsonResponse({"status": "error", "message": "No email address provided"})
    elif request.method == "GET":
        id = request.GET["id"]
        try:
            id = Otp.objects.using(DB_NAME).get(otp=id)
            current_time = timezone.make_aware(
                    datetime.datetime.utcnow().replace(microsecond=0))
            time_diff = current_time - id.created_date_time
            time_diff = time_diff.total_seconds()
            time_diff = round(time_diff / 60)
            print(time_diff)
            if time_diff <= 2880 and time_diff >= 0:
                user = id.user_id
                print("test -", user)
                id.delete()
            else:
                id.delete()
                raise ValueError(
                        "This session has been expired, Please Go back to <a href = " + LOGIN_URL + ">Login Page</a>")
        except ValueError as e:
            return HttpResponse(
                    "This session has been expired, Please Go back to <a href = " + LOGIN_URL + ">Login Page</a>")
        except Exception as e:
            print(e)
            return HttpResponse(
                    "This session has been expired, Please Go back to <a href = " + LOGIN_URL + ">http://eastlothian.tracsis-tads.com/aecon/eastlothian</a>")
        return render(request, 'aecon/reset-password.html',
                      {'background_image': "aecon/roadTraffic.jpg", 'user': user})


@login_required(login_url=LOGIN_URL)
def welcome_page(request):
    return render(request, 'aecon/welcome-page.html', {'background_image': "aecon/roadTraffic.jpg"})


# @login_required(login_url=LOGIN_URL)
def map_view(request):
    client = Client.objects.using(DB_NAME).get(id=18)
    return render(request, 'aecon/map-full-view.html', {"client": client})


# @login_required(login_url=LOGIN_URL)
def link_data(request, tmpltext):
    tmpltext = tmpltext.replace("/", "")
    print("tmpltext-", tmpltext)
    if tmpltext and request.user.is_authenticated:
        template = 'aecon/' + tmpltext + '.html'
        client = Client.objects.using(DB_NAME).get(users=request.user)
    elif request.user.is_authenticated:
        template = 'aecon/link-headata.html'
        tmpltext = "link-headata"
        client = Client.objects.using(DB_NAME).get(users=request.user)

    loc = client.locations.filter(
            observationType=10).order_by(Cast('name', CharField()))
    project = Project.objects.using(DB_NAME).all()
    return render(request, template,
                  {"client":   client, 'locations': loc, "mapCode": "cjwgah7js2kbb1cp9aztbp6vm", "flg": tmpltext,
                   "projects": project})


def check_if_still_processing(user):
    processing_path = os.path.join(EXCEL_ROOT, 'processing')
    json_file_path = os.path.join(processing_path, 'upload_data.json')

    if os.path.exists(json_file_path) and os.path.getsize(json_file_path) > 0:
        with open(json_file_path, 'r') as json_file:
            jsondata = json.load(json_file)
    else:
        jsondata = {}

    if not user.username in jsondata.keys():
        return

    for survey_type_flag in jsondata[user.username].keys():
        if not str(survey_type_flag) in jsondata[user.username]:
            return

        if jsondata[user.username][str(survey_type_flag)]['processing']:
            return

        if user.username in jsondata.keys():
            if str(survey_type_flag) in jsondata[user.username]:
                # Check if the files are still being processed
                upload_to_db_thread = threading.Thread(target=upload_data_to_db_thread,
                                                       args=(
                                                               user,
                                                               jsondata[user.username][str(survey_type_flag)]['files'],
                                                               int(survey_type_flag)))
                upload_to_db_thread.start()


# @login_required(login_url=LOGIN_URL)
def atc_data(request, tmpltext):
    print("hwllooo")
    tmpltext = tmpltext.replace("/", "")
    print("tmpltext-", tmpltext)
    if tmpltext and request.user.is_authenticated:
        template = 'aecon/' + tmpltext + '.html'
        print("template", template)
        client = Client.objects.using(DB_NAME).get(users=request.user)
    elif request.user.is_authenticated:
        print("hello111")
        template = 'aecon/atc-headline.html'
        tmpltext = "atc-headline"
        client = Client.objects.using(DB_NAME).get(users=request.user)
        print("client", client)
    else:
        print("hwloo33")
        template = 'aecon/atc-compare.html'
        tmpltext = "atc-compare"
        client = Client.objects.using(DB_NAME).get(id=18)
        print("client111", client)

    loc = client.locations.filter(
            observationType=9).order_by(Cast('name', CharField()))
    project = Project.objects.using(DB_NAME).all()
    return render(request, template,
                  {"client":   client, 'locations': loc, "mapCode": "cjwgah7js2kbb1cp9aztbp6vm", "flg": tmpltext,
                   "projects": project})


def radar_data(request, tmpltext):
    tmpltext = tmpltext.replace("/", "")
    print("tmpltext-", tmpltext)
    if tmpltext and request.user.is_authenticated:
        template = 'aecon/' + tmpltext + '.html'
        client = Client.objects.using(DB_NAME).get(users=request.user)
    elif request.user.is_authenticated:
        template = 'aecon/radar-headline.html'
        tmpltext = "radar-headline"
        client = Client.objects.using(DB_NAME).get(users=request.user)
    else:
        template = 'aecon/radar-compare.html'
        tmpltext = "radar-compare"
        client = Client.objects.using(DB_NAME).get(id=18)

    loc = client.locations.filter(observationType=12)

    projects = Project.objects.using(DB_NAME).all()
    return render(request, template,
                  {"client":   client, 'locations': loc, "mapCode": "cjwgah7js2kbb1cp9aztbp6vm", "flg": tmpltext,
                   "projects": projects})


@login_required(login_url=LOGIN_URL)
def sensor_data(request, tmpltext):
    tmpltext = tmpltext.replace("/", "")
    if tmpltext:
        template = 'aecon/' + tmpltext + '.html'
    else:
        template = 'aecon/sensor-headata.html'
        tmpltext = "sensor-headata"
    client = Client.objects.using(DB_NAME).get(users=request.user)
    loc = client.locations.filter(observationType=1)

    loc_data = {}
    for x in loc:
        try:
            area = x.area.name
        except Exception as e:
            area = None

        if area in loc_data.keys():
            loc_data[area].append(x)
        else:
            loc_data[area] = [x]

    return render(request, template,
                  {"client": client, "location_data": loc_data, "mapCode": "cjwgah7js2kbb1cp9aztbp6vm",
                   "flg":    tmpltext})


@login_required(login_url=LOGIN_URL)
def sensor_data_download(request):
    root_url = request.build_absolute_uri(f'/{BASE_URL}/')
    if request.method == "GET":
        if request.user.is_authenticated:
            client = Client.objects.using(DB_NAME).get(users=request.user)
        else:
            client = Client.objects.using(DB_NAME).get(id=18)
        loc = client.locations.filter(observationType=1)

        loc_data = {}
        for x in loc:
            try:
                area = x.area.name
            except Exception as e:
                area = None

            if area in loc_data.keys():
                loc_data[area].append(x)
            else:
                loc_data[area] = [x]
        template = 'aecon/download-view.html'
        return render(request, template, {"client": client, "location_data": loc_data})

    if request.method == "POST":
        params = request.POST.dict()
        template = 'aecon/download-view.html'

        try:
            params["startDate"] = datetime.datetime.strptime(
                    params["startDate"], "%Y-%m-%d")
            params["endDate"] = datetime.datetime.strptime(
                    params["endDate"], "%Y-%m-%d")
            params["classes"] = [int(c) for c in json.loads(params["classes"])]
            params["ids"] = json.loads(params["ids"])
            if request.user.is_authenticated:
                client = Client.objects.using(DB_NAME).get(users=request.user)
            else:
                client = Client.objects.using(DB_NAME).get(id=18)
            loc = client.locations.filter(observationType=1)

            loc_data = {}
            for x in loc:
                try:
                    area = x.area.name
                except Exception as e:
                    area = None

                if area in loc_data.keys():
                    loc_data[area].append(x)
                else:
                    loc_data[area] = [x]
            datapoints = calculate_num_datapoints(**params)
            if datapoints > 5000000:
                return HttpResponseBadRequest(
                        reason="You have requested too much data in a single request. Please break your request down into smaller downloads")
            if Location.objects.using(DB_NAME).filter(id__in=params["ids"]).values(
                    "observationType").distinct().count() > 1:
                return HttpResponseBadRequest(reason="You cannot download data for multiple different sensor types")
            assert client.locations.filter(
                    id__in=params["ids"]).count() == len(params["ids"])
            file = generate_salt(16) + str(datetime.datetime.now().date())
            params["clientTz"] = client.timezone
            params["table"] = request.session.session_key + "_download"
            params["dataType"] = "sensor"  # Not sure what to put here but needs to be defined to avoid error
            params["rootUrl"] = root_url

            if not params["email"]:
                params["email"] = request.user.email

            params["username"] = request.user.first_name

            threading.Thread(target=download_thread,
                             args=(file,), kwargs=params).start()
            if "email" in params and params["email"] != "":
                msg = "A link to your download will be sent to " + \
                      str(params["email"])
                status = "email"
            return render(request, template, {"client": client, "location_data": loc_data})
        except ValueError as e:
            return HttpResponseBadRequest(reason="Invalid or missing date")
        except KeyError as e:
            return HttpResponseBadRequest(reason="Missing field : " + str(e))
        except AssertionError as e:
            return HttpResponseBadRequest(reason="You do not have access to the data for one or more selected sites")


@login_required(login_url=LOGIN_URL)
def jtc_data(request, tmpltext):
    tmpltext = tmpltext.replace("/", "")

    if tmpltext and request.user.is_authenticated:
        template = 'aecon/' + tmpltext + '.html'
        client = Client.objects.using(DB_NAME).get(users=request.user)
    elif request.user.is_authenticated:
        template = 'aecon/jtc-headline.html'
        tmpltext = "jtc-headline"
        client = Client.objects.using(DB_NAME).get(users=request.user)
    else:
        print("hwloo33")
        template = 'aecon/jtc-headline.html'
        tmpltext = "jtc-headline"
        client = Client.objects.using(DB_NAME).get(id=18)
        print("client111", client)

    loc = client.locations.filter(
            observationType__in=(11, 13)).order_by(Cast('name', CharField()))
    project = Project.objects.using(DB_NAME).filter(survey_type='JTC')
    proj_locs = []
    for pro in project:
        temp = {
                'id':    pro.id,
                'name':  pro.project_no,
                'sites': [{'id': loc.id, 'name': loc.name} for loc in pro.location_set.all()]
        }
        proj_locs.append(temp)
    return render(request, template, {'locations': loc, 'projectsList': mark_safe(proj_locs), "flg": tmpltext})


# def historic_data(request, tmpltext):
#     tmpltext = tmpltext.replace("/", "")
#     print("tmpltext-", tmpltext)
#     if tmpltext and request.user.is_authenticated:
#         template = 'aecon/' + tmpltext + '.html'
#         client = Client.objects.using(DB_NAME).get(users=request.user)
#     elif request.user.is_authenticated:
#         template = 'aecon/historic-headline.html'
#         tmpltext = "historic-headline"
#         client = Client.objects.using(DB_NAME).get(users=request.user)
#
#     loc = client.locations.filter(observationType=10).order_by(Cast('api_identifier', IntegerField()))
#     project = Project.objects.using(DB_NAME).all()
#     return render(request, template,
#                   {"client": client, 'locations': loc, "mapCode": "cjwgah7js2kbb1cp9aztbp6vm", "flg": tmpltext,
#                    "projects": project})
#
# def mph_data(request, tmpltext):
#     tmpltext = tmpltext.replace("/", "")
#     print("tmpltext-", tmpltext)
#     if tmpltext and request.user.is_authenticated:
#         template = 'aecon/' + tmpltext + '.html'
#         client = Client.objects.using(DB_NAME).get(users=request.user)
#     elif request.user.is_authenticated:
#         template = 'aecon/mph-headline.html'
#         tmpltext = "mph-headline"
#         client = Client.objects.using(DB_NAME).get(users=request.user)
#
#     loc = client.locations.filter(observationType=1).order_by(Cast('api_identifier', IntegerField()))
#     project = Project.objects.using(DB_NAME).all()
#     return render(request, template,
#                   {"client": client, 'locations': loc, "mapCode": "cjwgah7js2kbb1cp9aztbp6vm", "flg": tmpltext,
#                    "projects": project})


class NpEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, np.integer):
            return int(obj)
        if isinstance(obj, np.floating):
            return float(obj)
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        return super(NpEncoder, self).default(obj)


@login_required(login_url=LOGIN_URL)
def get_atc_counts(request):
    params = request.POST.dict()
    ids = json.loads(request.POST["ids"])
    location = Location.objects.using(DB_NAME).filter(id__in=ids)
    project = location[0].projects.all()
    project = project[0]
    params['speed'] = True
    project_loc = ProjectLocations.objects.using(DB_NAME).get(
            project_id=params['project'], location=location[0])
    params['startDate'] = str(project_loc.startDate) + " 00:00"
    params['endDate'] = str(project_loc.endDate) + " 00:00"
    data = location.get_atc_counts(**params)
    graphLabels = []
    t1 = datetime.datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    t2 = datetime.datetime.now().replace(hour=23, minute=00, second=0, microsecond=0)
    while t2 > t1:
        graphLabels.append(t1.time().strftime("%H:%M"))
        t1 += datetime.timedelta(hours=1)
    graphLabels.append(t1.time().strftime("%H:%M"))

    # data = json.dumps(data, cls=NpEncoder)

    return JsonResponse({"status": "success", "data": data, "graphLabels": graphLabels})


@login_required(login_url=LOGIN_URL)
def get_atc_psl(request):
    params = request.POST.dict()
    location = Location.objects.using(DB_NAME).filter(id=params['id'])
    speed_limit = ProjectLocations.objects.using(DB_NAME).get(project_id=params['project'],
                                                              location_id=params['id']).speed_limit
    params['speed'] = True
    params['PSL'] = speed_limit
    params['ACPO'] = (1.1 * speed_limit) + 2
    params['DFT'] = speed_limit + 15
    data = location.get_atc_psl(**params)
    # print(speed_limit)
    return JsonResponse({"status": "success", "data": data})


# @login_required(login_url=LOGIN_URL)
def aggregated_data(request):
    loc_id = request.POST.get("location_id")

    speed_flg = request.POST.get("speed_flg")
    public_view = request.POST.get("public_view")
    loc = Location.objects.using(DB_NAME).get(id=loc_id)
    if speed_flg == 'false' and public_view == 'false':
        if not request.user.is_authenticated:
            return redirect(f"/{BASE_URL}")
        phase1 = request.POST.get("phase-header")
        phase2 = request.POST.get("phases-header")
        data = loc.bordersaggregateddata_set.filter(
                timeval__in=[8, 15, 24, 25, 26, 27], project__in=[phase1, phase2])
        print("got data")
        # print(data.query)
        # print(data)
        # for row in data:
        #     row.phase = 0 if row.project == phase1 else 1
        # print(data)
        phase1 = int(phase1)
        data = list(data.values("direction__order", "day", "timeval",
                                "avg", "perc_85th", "perc_95th", "project_id"))
        # print("converted in lists" , data)
        finalData = []
        for d in data:
            d['phase'] = 0 if (d['project_id'] == int(phase1)) else 1
        print("chnaged")
        return JsonResponse(data, safe=False)
    elif public_view == 'true':

        projects = list(
                ProjectLocations.objects.using(DB_NAME).filter(location=loc).order_by('startDate').values_list(
                        'project',
                        flat=True))

        data = loc.bordersaggregateddata_set.filter(
                timeval__in=[8, 15, 24, 25, 26, 27])
        # print((data.values("direction__direction__name")))
        data = list(data.values("direction__order", "day", "timeval",
                                "avg", "perc_85th", "perc_95th", "project_id"))
        for d in data:
            d['phase'] = projects.index(d['project_id'])
        return JsonResponse(data, safe=False)
    else:
        phase1 = request.POST.get("phase-header")
        phase2 = request.POST.get("phases-header")
        data = loc.bordersaggregateddata_set.filter(project__in=[phase1]).values("direction__order", "day", "timeval",
                                                                                 "avg", "perc_85th", "perc_95th")
        days = [x for x in range(1, 10)]
        for d in days:
            direction = [x for x in range(3)]
            for dir in direction:
                time = [x for x in range(28)]
                value = []
                for t in time:
                    for item in data:
                        if item['day'] == d and item['direction__order'] == dir and item['timeval'] == t:
                            value.append({"time":      t, 'avg': round(item['avg']),
                                          'perc_85th': round(item['perc_85th']),
                                          'perc_95th': round(item['perc_95th'])})
                direction[dir] = {'data': value}
            days[d - 1] = {'direction': direction}

        return JsonResponse(days, safe=False)


@login_required(login_url=LOGIN_URL)
def aggregated_headline_data(request):
    loc_id = request.POST.get("location_id")
    # client = Client.objects.using(DB_NAME).get(id=18)
    print(loc_id)
    loc = Location.objects.using(DB_NAME).get(id=loc_id)
    print(loc)
    data = BordersAggregatedData.objects.using(
            DB_NAME).filter(location=loc, day__in=[8, 9])

    data = list(
            data.values("project_id", "timeval", "direction__order", "direction__direction__descriptive", "day", "avg",
                        "perc_85th", "perc_95th", "counts"))

    result = sorted(data, key=operator.itemgetter("project_id"))
    outputList = {}

    for i, projects in itertools.groupby(result, key=operator.itemgetter("project_id")):
        projectsOutPut = {}
        projectsList = list([item for item in projects])

        projectsList = sorted(
                projectsList, key=operator.itemgetter("direction__order"))
        directionoutputList = {}
        for j, directions in itertools.groupby(projectsList, key=operator.itemgetter("direction__order")):
            directionsList = list([item for item in directions])
            directionsList = sorted(
                    directionsList, key=operator.itemgetter("day"))
            dayoutputList = {}
            for k, days in itertools.groupby(directionsList, key=operator.itemgetter("day")):
                dayLists = list([item for item in days])
                dayLists = sorted(dayLists, key=operator.itemgetter("timeval"))
                timeListsOutput = {}
                for l, items in itertools.groupby(dayLists, key=operator.itemgetter("timeval")):
                    itemLists = [{"avg":       round(item['avg']), "perc_85th": round(item["perc_85th"]),
                                  "perc_95th": round(item['perc_95th']),
                                  'direction': item['direction__direction__descriptive'],
                                  'counts':    round(item['counts'])} for item in items]
                    timeListsOutput[l] = itemLists[0]
                # print(timeListsOutput)
                dayoutputList[k] = timeListsOutput
            # print(dayoutputList)
            directionoutputList[j] = dayoutputList
        # print(projectsOutPut)
        outputList[i] = directionoutputList
    # print(outputList)
    result = outputList

    return JsonResponse(result, safe=False)


def download_survey(assoc):
    print("reached")
    wb = openpyxl.load_workbook(os.path.join(
            settings.BASE_DIR, "survey_download_template.xlsx"))
    sht = wb['Raw Data - To Be Deleted At End']
    print("wb -", sht.cell(row=10, column=1).value, sht.max_row)
    return
    for i in range(100):
        print("for i-", i)
        sht.cell(
                row=i + 10, column=1).value = assoc[i].date.strftime("%m/%d/%Y")
        sht.cell(row=i + 10, column=2).value = str(assoc[i].date.time())
        sht.cell(row=i + 10, column=3).value = assoc[i].direction.direction.abbrev + str(
                assoc[i].direction.order)
        sht.cell(row=i + 10, column=4).value = assoc[i].direction.direction.abbrev + str(
                assoc[i].direction.order)
    wb.save(os.path.join(settings.BASE_DIR, "data", "test.xlsx"))


def start_download_survey(request):
    loc_id = request.POST.get("location_id")
    project_id = request.POST.get("project_id")
    print("loc_id-", loc_id, project_id)
    assoc = AssociatedObservation.objects.using(DB_NAME).filter(
            location_id=loc_id, project_id=project_id)
    th = threading.Thread(target=download_survey, args=(assoc,)).start()
    response = {"status": "200", "msg": "Download file in few minutes"}
    return JsonResponse(response, safe=False)


# --------------------------------- UPLOAD SECTION ----------------------------------------

def upload_to_dir_thread(uploaded_files):
    fs = FileSystemStorage()
    processing_path = os.path.join(EXCEL_ROOT, 'processing')

    # Create the processing directory if it doesn't exist
    if not os.path.exists(processing_path):
        os.makedirs(processing_path)

    for file_obj in uploaded_files:
        file_path = os.path.join(processing_path, file_obj.name)

        # Write the file to the directory
        with fs.open(file_path, 'wb+') as destination:
            for chunk in file_obj.chunks():
                destination.write(chunk)

        # Read the file content and save it to the database
        with open(file_path, 'rb') as file:
            file_content = file.read()
            uploaded_file = UploadedFile(
                    file_name=file_obj.name,
                    file_content=file_content
            )
            uploaded_file.save()
    return


def debug_log_view(request):
    if not request.user.is_superuser:
        return HttpResponseForbidden("You are not authorized to view this page")
    log_file_path = os.path.join(settings.BASE_DIR, 'debug_output.log')
    with open(log_file_path, 'r') as log_file:
        log_content = log_file.read()
    return HttpResponse(f"<pre>{log_content}</pre>")


def file_explorer_view(request, path=''):
    if not request.user.is_superuser:
        return HttpResponseForbidden("You are not authorized to view this page")

    base_dir = settings.BASE_DIR
    abs_path = os.path.join(base_dir, path)

    if not os.path.exists(abs_path):
        raise Http404("Path not found")

    if os.path.isfile(abs_path):
        with open(abs_path, 'r') as file:
            return HttpResponse("<pre>{}</pre>".format(file.read()), content_type='text/html')

    directories = []
    files = []
    for entry in os.listdir(abs_path):
        if os.path.isdir(os.path.join(abs_path, entry)):
            directories.append(entry)
        else:
            files.append(entry)

    context = {
            'directories': directories,
            'files':       files,
            'path':        path,
    }
    return render(request, 'aecon/file_explorer.html', context)


def upload_data_to_db_thread(user, uploaded_files_names, survey_type_flag):
    print(f"Survey: {survey_type_flag}\nUploaded Files: {uploaded_files_names}")
    start_timeout_job()
    warnings.filterwarnings('ignore')
    processing_path = os.path.join(EXCEL_ROOT, 'processing')
    email = str(user.email)
    files_needed = []
    json_file_path = os.path.join(processing_path, 'upload_data.json')
    try:
        json_entity = JSONFile.objects.get(file_name='upload_data.json')
    except:
        print("No upload_data.json was found in the Database")
        json_entity = JSONFile(file_name='upload_data.json', file_content={})
        json_entity.save()

    # Read or create the JSON file
    if os.path.exists(json_file_path) and os.path.getsize(json_file_path) > 0:
        with open(json_file_path, 'r') as json_file:
            jsondata = json.load(json_file)
    else:
        jsondata = {}

    # Update the JSON file with the new information
    user_key = user.username
    if user_key not in jsondata:
        jsondata[user_key] = {}

    if str(survey_type_flag) not in jsondata[user_key]:
        jsondata[user_key][str(survey_type_flag)] = {}

    # Only set 'original files' if it doesn't exist
    if 'original files' not in jsondata[user_key][str(survey_type_flag)]:
        jsondata[user_key][str(survey_type_flag)
        ]['original files'] = uploaded_files_names.copy()

    jsondata[user_key][str(survey_type_flag)]['files'] = uploaded_files_names
    jsondata[user_key][str(survey_type_flag)]['processing'] = True
    jsondata[user_key][str(survey_type_flag)
    ]['process start time'] = time.time()

    # Write the updated JSON data back to the file
    with open(json_file_path, 'w') as json_file:
        json.dump(jsondata, json_file, indent=4)
        json_entity.file_content = jsondata
        json_entity.save()

    failed_file_errors = {}

    # Lists out the paths for all the files in the project
    for subdir, dirs, files in os.walk(processing_path):
        for file_name in files:
            if file_name in uploaded_files_names:
                filepath = os.path.join(subdir, file_name)
                files_needed.append(filepath)

    # Site Details
    for file_path in files_needed:
        print("aaaaaaaa", file_path)

        filename = os.path.basename(file_path)

        log = UploadLogging.objects.create(file_name=filename)
        site_filename = os.path.basename(file_path) + "Raw_sites.csv"
        raw_data_filename = os.path.basename(file_path) + "Raw_Data.csv"

        try:
            project_list = []
            log.converstion_to_CSV_start = timezone.now()
            log.save()

            if survey_type_flag == 10:
                jtc_df = pd.read_csv(file_path)
                project_number = jtc_df["Project No"][0]
                jtc_data_filename = os.path.basename(
                        file_path) + "Raw_JTC_Data.csv"
                jtc_df.to_csv(os.path.join(processing_path, jtc_data_filename))
                raw_data_filename = jtc_data_filename
                project_list.append(project_number)
            else:
                xlsx = pd.ExcelFile(file_path)
                print(xlsx.sheet_names)
                df1 = xlsx.parse('site_for_DB')
                df2 = xlsx.parse('Raw Data')

                project_number = df1["Project_No"][0]
                secondary_dir = df1["Secondary"][0]
                if isinstance(secondary_dir, (float, np.float64)):
                    df1["Secondary"][0] = 'NoDirection'

                site_filename = os.path.basename(file_path) + "Raw_sites.csv"
                raw_data_filename = os.path.basename(
                        file_path) + "Raw_Data.csv"

                df1.to_csv(os.path.join(processing_path, site_filename))
                df2.to_csv(os.path.join(processing_path, raw_data_filename))

                project_list.append(project_number)

            # Uploading to Database Part
            data_file_path = os.path.join(processing_path, raw_data_filename)
            sites_file_path = os.path.join(processing_path, site_filename)

            site = pd.read_csv(sites_file_path)
            site_data = site.iloc[0]
            site_no = site_data['Site_No']
            data = pd.read_csv(data_file_path)
            log.converstion_to_CSV_end = timezone.now()
            log.save()

            if survey_type_flag == 11:
                for index, row in data.iterrows():
                    data.at[index, 'Site_No'] = site_no

                    try:
                        # Try to parse the date in the 'year-month-day' format
                        datetime.datetime.strptime(row['Date'], '%Y-%m-%d')
                    except ValueError:
                        try:
                            correct_date = datetime.datetime.strptime(
                                    row['Date'], '%d/%m/%Y').strftime('%Y-%m-%d')
                            data.at[index, 'Date'] = correct_date
                        except ValueError:
                            print(
                                    f"Date {row['Date']} is not in expected format. Skipping.")

                # Write the DataFrame back to the CSV file
                data.to_csv(data_file_path, index=False)

            if survey_type_flag == 10:
                log.observation_upload_start = timezone.now()
                log.save()
                upload_jtc_raw_data_to_db(data_file_path, log)
            else:
                upload_site_information_to_db(
                        sites_file_path, survey_type_flag)
                upload_site_observations_to_db(
                        data_file_path, sites_file_path, survey_type_flag, log)

        except Exception as error:
            file_name = os.path.basename(file_path)
            log.error_txt = error
            log.save()
            if file_name not in failed_file_errors:
                failed_file_errors[file_name] = str(error)
                if 'failed files' not in jsondata[user_key][str(survey_type_flag)]:
                    jsondata[user_key][str(survey_type_flag)]['failed files'] = {
                            file_name: str(error)}
                else:
                    jsondata[user_key][str(
                            survey_type_flag)]['failed files'][file_name] = str(error)
                # Update JSON to remove the failed file
                if file_name in jsondata[user_key][str(survey_type_flag)]['files']:
                    jsondata[user_key][str(survey_type_flag)]['files'].remove(
                            file_name)

        try:
            file_to_delete = os.path.basename(file_path)
            UploadedFile.objects.get(file_name=file_to_delete).delete()
            os.unlink(file_path)
            os.unlink(os.path.join(processing_path, site_filename))
            os.unlink(os.path.join(processing_path, raw_data_filename))
            if file_to_delete in jsondata[user_key][str(survey_type_flag)]['files']:
                jsondata[user_key][str(survey_type_flag)
                ]['files'].remove(file_to_delete)

        except Exception as e:
            print(f'Failed to delete {file_path}. Reason: {e}')
            log.error_txt = f"failed to delete because {e}"
            log.save()

        try:
            with open(json_file_path, 'w') as json_file:
                json.dump(jsondata, json_file, indent=4)
                json_entity.file_content = jsondata
                json_entity.save()
        except Exception as e:
            print(f'Failed to write to {json_file_path}')
            log.error_txt = f"failed to delete because {e}"
            log.save()

        # Was for debugging purposes, Emailed the user (if admin) for each file completion.
        # if user.is_superuser:
        #     content = f"<p>Dear {user.first_name}, this file has just finished: {os.path.basename(file_path)}</p>"
        #     sendMail(email, "Upload Completed", content)

        log.process_end = timezone.now()
        log.save()

    if 'failed files' in jsondata[user_key][str(survey_type_flag)]:
        failed_file_errors = jsondata[user_key][str(
                survey_type_flag)]['failed files']

    original_files = jsondata[user_key][str(
            survey_type_flag)]['original files']

    content = "<html>"
    content += f"<p>Dear {user.first_name},</p>"
    content += "<br><p>Data Upload for AECON has been completed" + \
               (" with errors.</p>" if failed_file_errors else ".</p>")
    content += "<p>The Files uploaded are: " + ", ".join(os.path.basename(
            file) for file in original_files if os.path.basename(file) not in failed_file_errors.keys()) + "</p>"
    if failed_file_errors:
        content += "<p><strong>However these files could not be uploaded: </strong>" + \
                   ",<br> ".join(
                           file for file in failed_file_errors.keys()) + "<br></p>"
        content += "<p>More Info: " + \
                   ",<br> ".join(
                           f"<br>File Name:<br>&emsp;&emsp;{file}<br>File Error:<br>&emsp;&emsp;{error}" for file, error
                           in failed_file_errors.items()) + "<br></p>"
    content += "<p>You can now review it on the website.</p>"
    content += "<br><p>Kind Regards,<br>AECON Team</p>"
    content += "</hmtl>"

    # Remove any empty structures in the JSON file
    if not jsondata[user_key][str(survey_type_flag)]['files']:
        jsondata[user_key].pop(str(survey_type_flag))

    if not jsondata[user_key]:
        jsondata.pop(user_key)

    if not jsondata:
        print("Stoping check for unprocessed uploads shceduler job")
        stop_timeout_job()

    # Write the cleaned JSON data back to the file
    with open(json_file_path, 'w') as json_file:
        json.dump(jsondata, json_file, indent=4)
        json_entity.file_content = jsondata
        json_entity.save()

    try:
        sendMail(email, "Upload Completed", content)
    except Exception as error:
        print(f"Couldn't send email to {email}. Error occurred: {error}")

    return


# ----------------------------------------- DATA UPLOAD ---------------------------------------------------

def upload_site_observations_to_db(data_filename, site_filename, data_type_flag, log):
    log.observation_upload_start = timezone.now()
    log.save()
    warnings.filterwarnings('ignore')
    sites = pd.read_csv(site_filename)
    data_file = pd.read_csv(data_filename)
    for idx, row in data_file.iterrows():
        print(row["Site_No"])

    sites = sites.dropna()
    sites_list_of_dicts = []
    projects_list = []

    for index, row in sites.iterrows():
        test_project = Project.objects.using(DB_NAME).get(
                project_no=row['Project_No'].strip())
        test_location = Location.objects.using(DB_NAME).get(api_identifier=row['Site_No'],
                                                            observationType_id=data_type_flag)
        print("test location", test_location, "test project", test_project)
        if data_type_flag == 11:
            observations_exists = LINKObservation.objects.using(DB_NAME).filter(location=test_location,
                                                                                project=test_project)
        else:
            observations_exists = AssociatedObservation.objects.using(DB_NAME).filter(location=test_location,
                                                                                      project=test_project)
        if not observations_exists:
            print("will use bulk create method")
            sites_dict = {
                    "location":   test_location,
                    "project":    test_project,
                    "data_exist": False
            }
        else:
            print("again will use bulk create method")
            # delete from database
            try:
                log.existing_data_delete_start = timezone.now()
                log.save()
                observations_exists.delete()
                log.existing_data_delete_end = timezone.now()
                log.save()
            except Exception as e:
                print("Error deleting obsexi from database:", str(e))
                log.error_txt = e
                log.save()

            # try:
            #     test_project.delete()
            # except Exception as e:
            #     print("Error deleting proj from database:", str(e))

            # try:
            #     test_location.delete()
            # except Exception as e:
            #     print("Error deleting loc from database:", str(e))

            sites_dict = {
                    "location":   test_location,
                    "project":    test_project,
                    "data_exist": True
            }

        sites_list_of_dicts.append(sites_dict)
        projects_list.append(row['Project_No'])
    if data_type_flag == 11:
        class_lst = {1: 43107, 2: 43108, 3: 43109, 4: 43110, 5: 43111,
                     6: 43112, 7: 43113, 8: 43114, 9: 43115, 10: 43116, 11: 43117, 12: 43118}
    elif data_type_flag == 9:
        class_lst = {1:  43107, 2: 43108, 3: 43109, 4: 43110, 5: 43111, 6: 43112, 7: 43117, 8: 43118, 9: 43119,
                     10: 43120,
                     11: 43121, 12: 43122, 13: 43122}
    else:
        class_lst = {1:  43103, 2: 43104, 3: 43104, 4: 43105, 5: 43105, 6: 43106, 7: 43106, 8: 43106, 9: 43106,
                     10: 43106,
                     11: 43106, 12: 43106, 13: 43106}

    data = pd.read_csv(data_filename)
    if data_type_flag == 11:
        data = data.drop(['Note'], axis=1)
    data = data.dropna()
    data['Date'] = data['Date'].apply(lambda date: str(date.split(" ")[0]))

    # other format that i might need: 1- %m/%d/%Y 2-%Y-%m-%d
    data['Date'] = pd.to_datetime(data['Date'], format="%Y-%m-%d")
    data = data.fillna('')
    out = []

    if len(data.index) > 50000:
        for i in range(0, len(data.index), 50000):
            sub_data = pd.DataFrame()
            sub_data = data[i:i + 50000]
            out = []
            for index, row in sub_data.iterrows():
                Datetime = parser.parse(
                        str(row['Date']) + " " + str(row['Time']))

                location = Location.objects.using(DB_NAME).get(api_identifier=row['Site_No'],
                                                               observationType_id=data_type_flag)
                project = Project.objects.using(DB_NAME).get(
                        project_no=row['Project_No'].strip())
                if data_type_flag == 11:
                    weather_code = row['Weather_Code']
                    day_no = row['Day_No']
                    time_15min = row['Time 15min']
                    time_1hr = row['Time 1hr']
                    value = row['Volume']
                else:
                    value = row['Speed']

                idx = row['Index']
                tmp_dir = row['Direction'].strip()[0]

                DirectionList = location.directions.all()
                ClassList = location.classes.all()

                loc_class = ClassList.get(
                        location=location, obsClass_id=class_lst[row['Class']])

                # other implementation location=location, direction__abbrev=tmp_dir
                loc_dir = DirectionList.filter(
                        direction__abbrev=tmp_dir, ).first()

                data_exist_flag = list(filter(lambda obj: obj["location"] == location and obj["project"] == project,
                                              sites_list_of_dicts))

                if not data_exist_flag[0]['data_exist']:
                    if data_type_flag == 11:
                        assoc_observation = LINKObservation(location=location, date=Datetime, direction=loc_dir,
                                                            obsClass=loc_class, project=project, value=value,
                                                            obs_idx=idx,
                                                            weather_code=weather_code, day_no=day_no,
                                                            time_15min=time_15min,
                                                            time_1hr=time_1hr)
                    else:
                        assoc_observation = AssociatedObservation(location=location, date=Datetime, direction=loc_dir,
                                                                  obsClass=loc_class, project=project, value=value,
                                                                  obs_idx=idx, is_aggregated=False)

                    out.append(assoc_observation)
                    print("added to list")

                else:
                    print("adddddddddddddddddd")
                    if data_type_flag == 11:

                        assoc_observation = LINKObservation(location=location, date=Datetime, direction=loc_dir,
                                                            obsClass=loc_class, project=project, value=value,
                                                            obs_idx=idx,
                                                            weather_code=weather_code, day_no=day_no,
                                                            time_15min=time_15min,
                                                            time_1hr=time_1hr)
                        # LINKObservation.objects.using(DB_NAME).update_or_create(location=location,
                        #                                                         date=Datetime,
                        #                                                         direction=loc_dir,
                        #                                                         obsClass=loc_class,
                        #                                                         project=project, value=value,
                        #                                                         obs_idx=idx,
                        #                                                         defaults={'location': location,
                        #                                                                   'value': value,
                        #                                                                   'obsClass': loc_class,
                        #                                                                   'date': Datetime,
                        #                                                                   'direction': loc_dir,
                        #                                                                   'project': project,
                        #                                                                   'obs_idx': idx,
                        #                                                                   'weather_code': weather_code,
                        #                                                                   'day_no': day_no,
                        #                                                                   'time_15min': time_15min,
                        #                                                                   'time_1hr': time_1hr})
                    else:
                        print("aaaaaaaaaaaaaaaaaaaaaaaaaaaa")

                    assoc_observation = AssociatedObservation(location=location, date=Datetime, direction=loc_dir,
                                                              obsClass=loc_class, project=project, value=value,
                                                              obs_idx=idx, is_aggregated=False)
                    out.append(assoc_observation)

                    # AssociatedObservation.objects.using(DB_NAME).update_or_create(location=location,
                    #                                                               date=Datetime,
                    #                                                               direction=loc_dir,
                    #                                                               obsClass=loc_class,
                    #                                                               project=project, value=value,
                    #                                                               obs_idx=idx,
                    #                                                               defaults={'location': location,
                    #                                                                         'value': value,
                    #                                                                         'obsClass': loc_class,
                    #                                                                         'date': Datetime,
                    #                                                                         'direction': loc_dir,
                    #                                                                         'project': project,
                    #                                                                         'is_aggregated': False,
                    #                                                                         'obs_idx': idx})
            if out:
                print("successfully appended observations")
                try:
                    print("adding ", len(out), "observations")
                    if data_type_flag == 11:
                        LINKObservation.objects.using(DB_NAME).bulk_create(out, batch_size=1000,
                                                                           ignore_conflicts=False)
                    else:
                        AssociatedObservation.objects.using(DB_NAME).bulk_create(out, batch_size=1000,
                                                                                 ignore_conflicts=False)
                    print("successfully added observations")
                except django.db.IntegrityError as e:
                    print("counts were already in the database")
                except Exception as e:
                    print(e)
                    print("something went wrong when adding observations")
                    raise django.db.Error(
                            "Failed to write data to database " + str(e))
            else:
                print(
                        "Data Already Exists for this project and location - Used update_or_create method")

    else:
        for index, row in data.iterrows():
            print("hhhhhhhhhhhhhhh", row['Site_No'])
            print(row["Date"])
            Datetime = parser.parse(str(row['Date']) + " " + str(row['Time']))
            location = Location.objects.using(DB_NAME).get(api_identifier=row['Site_No'],
                                                           observationType_id=data_type_flag)
            print("xxxxxxxxx", location)

            project = Project.objects.using(DB_NAME).get(
                    project_no=row['Project_No'].strip())

            if data_type_flag == 11:
                weather_code = row['Weather_Code']
                day_no = row['Day_No']
                time_15min = row['Time 15min']
                time_1hr = row['Time 1hr']
                value = row['Volume']
            else:
                value = row['Speed']

            idx = row['Index']

            tmp_dir = row['Direction'].strip()[0]
            DirectionList = location.directions.all()
            ClassList = location.classes.all()
            # loc_class = ClassList.get(
            #     location=location, obsClass_id=class_lst[row['Class']])
            print(row["Class"])
            print(class_lst)
            loc_class = LocationObservationClass.objects.using(DB_NAME).get(
                    location=location, obsClass_id=class_lst[row['Class']])
            print("hii", class_lst[row['Class']])

            # other implementation location=location, direction__abbrev=tmp_dir
            loc_dir = DirectionList.filter(direction__abbrev=tmp_dir, ).first()
            # print("aaaaaaaa", loc_dir, day_no,weather_code)

            flagg = list(filter(lambda obj: obj["location"] == location and obj["project"] == project,
                                sites_list_of_dicts))
            print(flagg)

            if not flagg[0]['data_exist']:
                print("xxxxxxxxxxxxxxxxxxxx")
                if data_type_flag == 11:
                    assoc_observation = LINKObservation(location=location, date=Datetime, direction=loc_dir,
                                                        obsClass=loc_class, project=project, value=value, obs_idx=idx,
                                                        weather_code=weather_code, day_no=day_no, time_15min=time_15min,
                                                        time_1hr=time_1hr)
                else:
                    assoc_observation = AssociatedObservation(location=location, date=Datetime, direction=loc_dir,
                                                              obsClass=loc_class, project=project, value=value,
                                                              obs_idx=idx, is_aggregated=False)
                out.append(assoc_observation)
                print("added observation - ", index)
            else:
                print("fffffffffff")
                if data_type_flag == 11:
                    if i == 0:
                        log.existing_data_delete_start = timezone.now()
                        log.save()
                        LINKObservation.objects.using(DB_NAME).filter(location=location,
                                                                      obsClass=loc_class,
                                                                      project=project).delete()
                        log.existing_data_delete_end = timezone.now()
                        log.save()

                        assoc_observation = LINKObservation(location=location, date=Datetime, direction=loc_dir,
                                                            obsClass=loc_class, project=project, value=value,
                                                            obs_idx=idx,
                                                            weather_code=weather_code, day_no=day_no,
                                                            time_15min=time_15min,
                                                            time_1hr=time_1hr)
                        # LINKObservation.objects.using(DB_NAME).update_or_create(location=location,
                        #                                                         date=Datetime,
                        #                                                         direction=loc_dir,
                        #                                                         obsClass=loc_class,
                        #                                                         project=project, value=value,
                        #                                                         obs_idx=idx,
                        #                                                         defaults={'location': location,
                        #                                                                   'value': value,
                        #                                                                   'obsClass': loc_class,
                        #                                                                   'date': Datetime,
                        #                                                                   'direction': loc_dir,
                        #                                                                   'project': project,
                        #                                                                   'obs_idx': idx,
                        #                                                                   'weather_code': weather_code,
                        #                                                                   'day_no': day_no,
                        #                                                                   'time_15min': time_15min,
                        #                                                                   'time_1hr': time_1hr})
                else:
                    print("aaaaaaaaaaaaaaaaaaaaaaaaaaaa")
                    if index == 0:
                        log.existing_data_delete_start = timezone.now()
                        log.save()
                        AssociatedObservation.objects.using(DB_NAME).filter(location=location, obsClass=loc_class,
                                                                            project=project).delete()
                        log.existing_data_delete_end = timezone.now()
                        log.save()

                    assoc_observation = AssociatedObservation(location=location, date=Datetime, direction=loc_dir,
                                                              obsClass=loc_class, project=project, value=value,
                                                              obs_idx=idx, is_aggregated=False)
                    out.append(assoc_observation)

                    # AssociatedObservation.objects.using(DB_NAME).update_or_create(location=location,
                    #                                                               date=Datetime,
                    #                                                               direction=loc_dir,
                    #                                                               obsClass=loc_class,
                    #                                                               project=project, value=value,
                    #                                                               obs_idx=idx,
                    #                                                               defaults={'location': location,
                    #                                                                         'value': value,
                    #                                                                         'obsClass': loc_class,
                    #                                                                         'date': Datetime,
                    #                                                                         'direction': loc_dir,
                    #                                                                         'project': project,
                    #                                                                         'is_aggregated': False,
                    #                                                                         'obs_idx': idx})
                print("added observations -", index)

        if out:
            print("successfully appended observations")
            try:
                print("adding ", len(out), "observations")
                if data_type_flag == 11:
                    LINKObservation.objects.using(DB_NAME).bulk_create(out, batch_size=1000,
                                                                       ignore_conflicts=False)
                else:
                    AssociatedObservation.objects.using(DB_NAME).bulk_create(
                            out, batch_size=1000, ignore_conflicts=False)
                print("successfully added observations")
            except django.db.IntegrityError as e:
                print("counts were already in the database")
            except Exception as e:
                print(e)
                print("something went wrong when adding observations")
                raise django.db.Error(
                        "Failed to write data to database " + str(e))
        else:
            print(
                    "Data Already Exists for this project and location - Used update_or_create method")
    log.observation_upload_end = timezone.now()
    log.save()
    if data_type_flag != 11:
        log.aggregation_start = timezone.now()
        log.save()
        for project_no in set(projects_list):
            update_aggregated_data(project_no, data_type_flag, log)


def update_aggregated_data(project_no, data_type_flag, log, loc_id=None):
    # get all data for a project as it will always calculate for a project.
    sys.setrecursionlimit(10000)
    project = Project.objects.using(DB_NAME).get(project_no=project_no)
    if loc_id != None:
        locations = Location.objects.using(DB_NAME).filter(projects=project, id=loc_id,
                                                           observationType_id=data_type_flag).prefetch_related()
    else:
        locations = Location.objects.using(DB_NAME).filter(projects=project,
                                                           observationType_id=data_type_flag).prefetch_related()

    for loc in locations:
        observation_data = AssociatedObservation.objects.using(DB_NAME).filter(location=loc, project=project,
                                                                               is_aggregated=False)
        if not observation_data:
            continue
        else:
            df = pd.DataFrame(list(observation_data.values()))
            df['dayofweek'] = df['date'].dt.dayofweek + 1
            df['hour'] = df['date'].dt.hour

            df.drop(['status', 'removed', 'date',
                     'obsClass_id', 'id'], axis=1, inplace=True)

            days = [1, 2, 3, 4, 5, 6, 7, 8, 9]
            dirs = loc.ordered_location_direction_queryset()
            for dir in dirs:

                dir_in = [dir.id]
                if dir.order != 2:
                    filtered_dir_df = df[(df['direction_id'] == dir.id)]
                else:
                    filtered_dir_df = df
                for day in days:
                    min_day = day
                    max_day = day
                    if day < 8:
                        filtered_day_df = filtered_dir_df[(
                                filtered_dir_df['dayofweek'] == day)]
                    else:
                        if day == 8:
                            min_day = 1
                            max_day = 5
                            filtered_day_df = filtered_dir_df[
                                (filtered_dir_df['dayofweek'] >= min_day) & (filtered_dir_df['dayofweek'] <= max_day)]
                        if day == 9:
                            filtered_day_df = filtered_dir_df
                    for timeval in range(0, 28):
                        time_min = timeval
                        time_max = timeval

                        if timeval < 24:
                            filtered_df = filtered_day_df[(
                                    filtered_day_df['hour'] == timeval)]

                        else:
                            if timeval == 24:
                                time_min = 7
                                time_max = 19
                            if timeval == 25:
                                time_min = 6
                                time_max = 22
                            if timeval == 26:
                                time_min = 6
                                time_max = 24
                            if timeval == 27:
                                time_min = 0
                                time_max = 24
                            filtered_df = filtered_day_df[
                                (filtered_day_df['hour'] >= time_min) & (filtered_day_df['hour'] <= time_max)]

                        avg = filtered_df['value'].mean(skipna=True)
                        perc_85th = filtered_df['value'].quantile(0.85)
                        perc_95th = filtered_df['value'].quantile(0.95)

                        filtered_df = filtered_df.groupby(
                                ['location_id', 'project_id', 'dayofweek']).size().reset_index(name='value')

                        filtered_df = filtered_df.groupby(['location_id', 'project_id'], as_index=False).agg(
                                {"value": lambda x: x.mean(skipna=True)})

                        counts = 0
                        if len(filtered_df.index) > 1:
                            print(filtered_df)
                            raise ValueError("count check")
                        elif len(filtered_df.index) == 0:
                            counts = 0
                        else:
                            counts = filtered_df['value'].iloc[0]

                        avg = round(avg, 2) if not np.isnan(float(avg)) else 0
                        perc_85th = round(perc_85th, 2) if not np.isnan(
                                float(perc_85th)) else 0
                        perc_95th = round(perc_95th, 2) if not np.isnan(
                                float(perc_95th)) else 0

                        print(avg, perc_85th, perc_95th, counts)

                        BordersAggregatedData.objects.using(DB_NAME).update_or_create(location=loc,
                                                                                      day=day, direction=dir,
                                                                                      project=project, timeval=timeval,
                                                                                      defaults={'location':  loc,
                                                                                                'day':       day,
                                                                                                'direction': dir,
                                                                                                'project':   project,
                                                                                                'timeval':   timeval,
                                                                                                'avg':       avg,
                                                                                                'perc_85th': perc_85th,
                                                                                                'perc_95th': perc_95th,
                                                                                                'phase':     1,
                                                                                                'counts':    counts})

            AssociatedObservation.objects.using(DB_NAME).filter(location=loc, project=project, is_aggregated=False) \
                .update(is_aggregated=True)
    log.aggregation_end = timezone.now()
    log.save()


# for importing locations
def upload_site_information_to_db(filename, data_type_flag):
    data = pd.read_csv(filename)
    data = data.fillna("")

    print(data)
    if data_type_flag == 9:
        class_id = [43107, 43108, 43109, 43110, 43111,
                    43112, 43117, 43118, 43119, 43120, 43121, 43122]
    elif data_type_flag == 12:
        class_id = [43137, 43138, 43139, 43140, 43141]

    for index, row in data.iterrows():
        # for adding Project
        project = Project.objects.using(DB_NAME).update_or_create(project_no=row['Project_No'].strip(),
                                                                  name=row['Project_No'].strip(
                                                                  ),
                                                                  defaults={'survey_type': row['Survey_Type'].strip()})
        project = project[0]
        print(project)
        # -------------------------------------------------------------------------------------------

        # for adding Location
        primary_dir = Direction.objects.using(DB_NAME).get(
                descriptive=row['Primary'].strip())
        sec_dir = Direction.objects.using(DB_NAME).get(
                descriptive=row['Secondary'].strip())
        combined = Direction.objects.using(DB_NAME).get(id=15)
        dir_lst = [primary_dir, sec_dir, combined]

        location = Location.objects.using(DB_NAME).get_or_create(observationType_id=data_type_flag,
                                                                 api_identifier=row['Site_No'],
                                                                 defaults={'name':               row[
                                                                                                     'Site_Name'].strip(),
                                                                           'observationType_id': data_type_flag,
                                                                           'api_identifier':     row['Site_No'],
                                                                           'lat':                row['Lat'],
                                                                           'lon':                row['Long'],
                                                                           'speedLimit':         row[
                                                                                                     'Speed_Limit'] if data_type_flag == 9 or data_type_flag == 12 else None})
        location = location[0]

        # For adding the location to the client
        client = Client.objects.using(DB_NAME).get(id=18)
        client.locations.add(location)

        project = Project.objects.using(DB_NAME).get(
                project_no=row['Project_No'].strip())
        startDate = datetime.datetime.strptime(
                str(row['startDate']), "%Y-%m-%d").date()  # 1- %Y-%m-%d, 2- %m/%d/%Y
        endDate = datetime.datetime.strptime(
                str(row['endDate']), "%Y-%m-%d").date()

        ProjectLocations.objects.using(DB_NAME).update_or_create(location=location, project=project,
                                                                 defaults={'location':    location,
                                                                           'speed_limit': row[
                                                                                              'Speed_Limit'] if data_type_flag == 9 or data_type_flag == 12 else None,
                                                                           'project':     project,
                                                                           'startDate':   startDate,
                                                                           'endDate':     endDate})
        print("class_id", class_id)
        for i in range(len(class_id)):
            LocationObservationClass.objects.using(DB_NAME).get_or_create(location=location, obsClass_id=class_id[i],
                                                                          order=i, defaults={'location':    location,
                                                                                             'obsClass_id': class_id[i],
                                                                                             'order':       i})

        # ----------------------------------------------------------------------------------------------

        # for mapping of location direction

        # loc_dir = []
        for i in range(len(dir_lst)):
            print(dir_lst[i])
            print(location)
            print(i)
            # direction_id = dir_lst[i].id,
            LocationDirection.objects.using(DB_NAME).get_or_create(location=location, order=i,
                                                                   defaults={'location':     location,
                                                                             'direction_id': dir_lst[i].id,
                                                                             'order':        i})
            # loc_dir.append(temp[0])
        # ----------------------------------------------------------------------------------------------

        print("Added loc - ", location)
        print("-----------------------------------------------------")


@login_required(login_url=LOGIN_URL)
def atc_upload(request):
    if request.method == 'POST':
        current_user = request.user

        print("ATC Data Upload is Starting")

        uploaded_files = request.FILES.getlist('document')
        print("uploaded_files", uploaded_files)
        uploaded_files_list = []

        for file_obj in uploaded_files:
            uploaded_files_list.append(file_obj.name)

        pname = str(uploaded_files[0].name)
        pname = re.split('\s+|_', pname)[0]

        try:
            upload_to_dir = threading.Thread(
                    target=upload_to_dir_thread, args=(uploaded_files,))
            upload_to_dir.start()
            upload_to_dir.join()
            upload_to_db_thread = threading.Thread(target=upload_data_to_db_thread,
                                                   args=(current_user, uploaded_files_list, 9))
            upload_to_db_thread.start()
        except Exception as e:
            print("something happened during the ATC upload thread")

        messages.success(
                request, "The Project that you are uploading is: " + pname)

    return render(request, 'aecon/atc-upload.html')


@login_required(login_url=LOGIN_URL)
def radar_upload(request):
    if request.method == 'POST':

        current_user = request.user
        print("Radar Data Upload is Starting")

        uploaded_files = request.FILES.getlist('document')
        uploaded_files_list = []

        for file_obj in uploaded_files:
            uploaded_files_list.append(file_obj.name)

        pname = str(uploaded_files[0].name)
        pname = re.split('\s+|_', pname)[0]

        try:
            upload_to_dir = threading.Thread(
                    target=upload_to_dir_thread, args=(uploaded_files,))
            upload_to_dir.start()
            upload_to_dir.join()
            upload_to_db_thread = threading.Thread(target=upload_data_to_db_thread,
                                                   args=(current_user, uploaded_files_list, 12))
            upload_to_db_thread.start()
        except Exception as e:
            print("something happened during the Radar upload thread")

        messages.success(
                request, "The Project that you are uploading is: " + pname)

    return render(request, 'aecon/radar-upload.html')


@login_required(login_url=LOGIN_URL)
def jtc_upload(request):
    if request.method == 'POST':
        print("JTC Data Upload is Starting")
        token = generate_salt(16)

        uploaded_files = request.FILES.getlist('document')
        uploaded_files_list = []

        for file_obj in (uploaded_files):
            uploaded_files_list.append(file_obj.name)
        try:
            upload_to_dir = threading.Thread(
                    target=upload_to_dir_thread, args=(uploaded_files,))
            upload_to_dir.start()
            upload_to_dir.join()
            upload_to_db_thread = threading.Thread(target=upload_jtc_data_todb,
                                                   args=(request.user, uploaded_files_list, 11, token))
            upload_to_db_thread.start()
            {'status': 'sucess', 'token': token}
        except Exception as e:
            print("something happened during the ATC upload thread")
        return JsonResponse({'status': 'sucess', 'token': token})
    else:
        return render(request, 'aecon/jtc-upload.html')


@login_required(login_url=LOGIN_URL)
def jtc_turning_count_upload(request):
    if request.method == 'POST':
        print("JTC Data Upload is Starting")
        token = generate_salt(16)
        uploaded_files = request.FILES.getlist('document')
        uploaded_files_list = []

        for file_obj in (uploaded_files):
            uploaded_files_list.append(file_obj.name)
        try:
            upload_to_dir = threading.Thread(
                    target=upload_to_dir_thread, args=(uploaded_files,))
            upload_to_dir.start()
            upload_to_dir.join()

            upload_to_db_thread = threading.Thread(
                    target=turning_count_thread, args=(uploaded_files_list, token))
            upload_to_db_thread.start()
        except Exception as e:
            print("something happened during the ATC upload thread")
        return JsonResponse({'status': 'sucess', 'token': token})
    else:
        return render(request, 'aecon/jtc-turning-count-upload.html')


@login_required(login_url=LOGIN_URL)
def jtc_download(request):
    print(request.method)
    if request.method == 'GET':
        client = Client.objects.using(DB_NAME).get(id=18)
        loc = client.locations.filter(
                observationType=11).order_by(Cast('name', CharField()))
        projects = Project.objects.using(DB_NAME).filter(survey_type__in=['JTC', 'Turning Counts'])
        proj_locs = []
        for pro in projects:
            temp = {
                    'id':           pro.id,
                    'name':         pro.project_no,
                    'sites':        [{'id': loc.id, 'name': loc.name, 'obs_type': loc.observationType_id} for loc in
                                     pro.location_set.all()],
                    'project_type': pro.survey_type
            }
            proj_locs.append(temp)

        return render(request, 'aecon/jtc-download.html', {'project_locations': mark_safe(proj_locs)})
    elif request.method == 'POST':
        root_url = request.build_absolute_uri(f'/{BASE_URL}/')
        file = generate_salt(16) + str(datetime.datetime.now().date())
        response = {'status':  'sucess',
                    'message': 'Downloading process started'}
        try:
            params = request.POST.dict()
            project_ids = params['project_id'].split(',')
            location_ids = params['location_id'].split(',')
            is_turning_count = params['is_turning_count'] == 'true'
            threading.Thread(target=jtc_download_thread, args=(
                    request.user, project_ids, location_ids, root_url, file, is_turning_count)).start()
        except Exception as e:
            response = {'status': 'error', 'message': f'{e}'}

        return JsonResponse(response)


def jtc_download_thread(user, project_ids, location_ids, root_url, token, is_turning_count):
    try:
        file_path = os.path.join(EXCEL_ROOT, token)
        os.makedirs(EXCEL_ROOT, exist_ok=True)
        projects = Project.objects.filter(id__in=project_ids)
        excel_streams = {}

        extracted_value = ['project__project_no', 'project__name', 'location__api_identifier',
                           'location__lat', 'location__lon', 'start_time', 'end_time', 'peak_hour', 'origin_arm__name',
                           'destination_arm__name', 'obsClass__order', 'count', 'pcu']

        df_columns = ['Project No', 'Project Name', 'Site Number', 'Latitude', 'Longitude',
                      'Start_Time', 'End_Time', 'Peak Hour', 'Origin_Arm', 'Destination_Arm', 'Vechicle Class',
                      'Vechile Count', 'PCU Value']
        if is_turning_count:
            extracted_value.extend(['start_date', 'end_date'])
            df_columns.extend(['Start_Date', 'End_Date'])
        for proj in projects:
            locations = proj.location_set.filter(id__in=location_ids)
            for loc in locations:
                jtc_data = Jtc_Data.objects.filter(project__id__in=project_ids, location=loc).values_list(
                        *extracted_value)

                jtc_df = pd.DataFrame.from_records(jtc_data, columns=df_columns)

                if is_turning_count:
                    jtc_df['Start_Date'] = pd.to_datetime(
                            jtc_df.Start_Date.astype(str) + ' ' + jtc_df.Start_Time.astype(str))
                    jtc_df['End_Date'] = pd.to_datetime(jtc_df.End_Date.astype(str) + ' ' + jtc_df.End_Time.astype(str))

                    # insert column with insert(location, column_name, column_value)
                    jtc_df.insert(5, "Start_Date", jtc_df.pop("Start_Date"))
                    jtc_df.insert(6, "End_Date", jtc_df.pop("End_Date"))
                    jtc_df.pop("Start_Time")
                    jtc_df.pop("End_Time")

                # we are adding +1 to order of class as we uploading the classes in the same class_id -1
                jtc_df["Vechicle Class"] = jtc_df["Vechicle Class"] + 1

                survey_type = proj.survey_type
                project_no = proj.project_no
                project_name = proj.name
                site_id = f'{project_no}_{proj.survey_type}_{loc.api_identifier}'
                site_no = loc.api_identifier
                site_name = loc.name

                arms_data = Arms.objects.filter(location=loc, project=proj).extra(
                        select={'Site_No': site_no, 'Project_No': f"'{project_no}'"}).values_list('name', 'lat', 'lon',
                                                                                                  'Site_No',
                                                                                                  'Project_No',
                                                                                                  'display_name')

                arms_df = pd.DataFrame.from_records(arms_data, columns=[
                        'Name', 'Lat', 'Long', 'Site_No', 'Project_No', 'Display_Name'])

                site_data = ProjectLocations.objects.filter(project=proj, location=loc).extra(
                        select={'Site_ID':      F"'{site_id}'", 'Project_No': f"'{project_no}'",
                                'Project_Name': f"'{project_name}'", 'Site_No': f"'{site_no}'",
                                'Site_Name':    f"'{site_name}'",
                                'Survey_Type':  f"'{survey_type}'", 'Lat': loc.lat, 'Long': loc.lon}).values_list(
                    'Site_ID',
                    'Project_No',
                    'Project_Name',
                    'Site_No',
                    'Site_Name',
                    'Survey_Type',
                    'Lat',
                    'Long',
                    'startDate',
                    'endDate')

                site_df = pd.DataFrame.from_records(site_data, columns=[
                        'Site_ID', 'Project_No', 'Project_Name', 'Site_No', 'Site_Name', 'Survey_Type', 'Lat', 'Long',
                        'startDate', 'endDate'])

                with NamedTemporaryFile(delete=False) as filetmp:
                    with pd.ExcelWriter(filetmp, engine='xlsxwriter') as writer:
                        site_df.to_excel(
                                writer, sheet_name=site_id, index=False)
                        arms_df.to_excel(
                                writer, sheet_name='Arms', index=False)
                        jtc_df.to_excel(
                                writer, sheet_name='Raw_Data', index=False)
                    filetmp.seek(0)
                    filestream = filetmp.read()
                    filename = f'{proj.project_no}_{loc.name}.xlsx'
                    excel_streams[filename] = filestream
                print('Created excel file')

        with zipfile.ZipFile(file_path, 'w', zipfile.ZIP_DEFLATED) as archive:
            for name, stream in excel_streams.items():
                archive.writestr(zinfo_or_arcname=name, data=stream)

        formatted_date = datetime.datetime.now().strftime('%d-%m-%Y')

        name = f'Tracsis%20JTC%20{formatted_date}'
        url = f'{root_url}downloadFile?token={token}&name={name}&extension=zip'
        email = str(user.email)
        content = "<p>Dear " + str(user.first_name) + ",</p>"
        content += f"<p>The JTC Data you requested from AECON is ready. Please note that, <a href='{url}'> link to the data, this link is available for 15 days only.</a></p>"
        content += "<p>Please don’t hesitate to contact us, If you have any issue.</p>"
        content += "<p>Kind Regards,"
        content += "<br>AECON Team</p>"
        content += "<p>ttd.admin@tracsis.com</p>"
        sendMail(email, 'Data Download is Ready', content)
        print('Email Sent')
    except Exception as e:
        print('error -', e)
        mail_list = ['rohel@divyaltech.com']
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        sendMail(mail_list, 'Error in Downloading JTC Data',
                 f'Error-{e} , other info - {exc_type}, {fname}, {exc_tb.tb_lineno}')


@login_required(login_url=LOGIN_URL)
def link_upload(request):
    if request.method == 'POST':
        current_user = request.user

        print("LINK Data Upload is Starting")

        uploaded_files = request.FILES.getlist('document')
        uploaded_files_list = []

        for file_obj in uploaded_files:
            uploaded_files_list.append(file_obj.name)

        pname = str(uploaded_files[0].name)
        pname = re.split('\s+|_', pname)[0]

        try:
            upload_to_dir = threading.Thread(
                    target=upload_to_dir_thread, args=(uploaded_files,))
            upload_to_dir.start()
            upload_to_dir.join()
            upload_to_db_thread = threading.Thread(target=upload_data_to_db_thread,
                                                   args=(current_user, uploaded_files_list, 10))
            upload_to_db_thread.start()
        except Exception as e:
            print("something happened during the LINK upload thread")

        messages.success(
                request, "The Project that you are uploading is: " + pname)

    return render(request, 'aecon/link-upload.html')


@login_required(login_url=LOGIN_URL)
def link_data_download(request, survey_type='link'):
    if request.method == "POST":
        current_user = request.user
        project_location_list = request.POST.getlist('project_location')
        root_url = request.build_absolute_uri(f'/{BASE_URL}/')
        downloading_thread = threading.Thread(target=data_download_thread,
                                              args=(project_location_list, current_user, 10, root_url))
        downloading_thread.start()

        if request.user.is_authenticated:
            client = Client.objects.using(DB_NAME).get(users=request.user)
        else:
            client = Client.objects.using(DB_NAME).get(id=18)
        if survey_type:
            projects = Project.objects.filter(survey_type=survey_type)
        else:
            return HttpResponse(status=404)
        project_locations = {}
        for project in projects:
            project_locations[project] = project.location_set.all()
        template = 'aecon/link-data-download.html'
        return render(request, template,
                      {"client":      client, "projects": projects, "project_locations": project_locations,
                       "survey_type": survey_type})

    elif request.method == "GET":
        print(survey_type)
        if request.user.is_authenticated:
            client = Client.objects.using(DB_NAME).get(users=request.user)
        else:
            client = Client.objects.using(DB_NAME).get(id=18)
        if survey_type:
            projects = Project.objects.filter(survey_type=survey_type)
        else:
            return HttpResponse(status=404)
        project_locations = {}
        for project in projects:
            project_locations[project] = project.location_set.all()
        template = 'aecon/link-data-download.html'
        return render(request, template,
                      {"client":      client, "projects": projects, "project_locations": project_locations,
                       "survey_type": survey_type})


@login_required(login_url=LOGIN_URL)
def data_download(request, survey_type='atc'):
    print("Called :O ")
    if request.method == "POST":
        current_user = request.user
        project_location_list = request.POST.getlist('project_location')
        root_url = request.build_absolute_uri(f'/{BASE_URL}/')
        downloading_thread = threading.Thread(
                target=data_download_thread, args=(project_location_list, current_user, 9, root_url))
        downloading_thread.start()

        if request.user.is_authenticated:
            client = Client.objects.using(DB_NAME).get(users=request.user)
        else:
            client = Client.objects.using(DB_NAME).get(id=18)
        if survey_type == "atc" or survey_type == "link":
            projects = Project.objects.filter(survey_type=survey_type)
        else:
            return HttpResponse(status=404)
        project_locations = {}
        for project in projects:
            project_locations[project] = project.location_set.all()
        template = 'aecon/data-download.html'
        return render(request, template,
                      {"client":      client, "projects": projects, "project_locations": project_locations,
                       "survey_type": survey_type})

    elif request.method == "GET":
        if request.user.is_authenticated:
            client = Client.objects.using(DB_NAME).get(users=request.user)
        else:
            client = Client.objects.using(DB_NAME).get(id=18)
        if survey_type == "atc" or survey_type == "link":
            projects = Project.objects.filter(survey_type=survey_type)
        else:
            return HttpResponse(status=404)
        project_locations = {}
        for project in projects:
            project_locations[project] = project.location_set.all()
        template = 'aecon/data-download.html'
        return render(request, template,
                      {"client":      client, "projects": projects, "project_locations": project_locations,
                       "survey_type": survey_type})


@login_required(login_url=LOGIN_URL)
def radar_download(request, survey_type='radar'):
    root_url = request.build_absolute_uri(f'/{BASE_URL}/')
    if request.method == "POST":
        current_user = request.user
        project_location_list = request.POST.getlist('project_location')

        downloading_thread = threading.Thread(
                target=data_download_thread, args=(project_location_list, current_user, 12, root_url))
        downloading_thread.start()

        if request.user.is_authenticated:
            client = Client.objects.using(DB_NAME).get(users=request.user)
        else:
            client = Client.objects.using(DB_NAME).get(id=18)
        if survey_type:
            projects = Project.objects.filter(survey_type=survey_type)
        else:
            return HttpResponse(status=404)
        project_locations = {}
        for project in projects:
            project_locations[project] = project.location_set.all()
        template = 'aecon/radar-download.html'
        return render(request, template,
                      {"client":      client, "projects": projects, "project_locations": project_locations,
                       "survey_type": survey_type})
    elif request.method == "GET":
        print(survey_type)
        if request.user.is_authenticated:
            client = Client.objects.using(DB_NAME).get(users=request.user)
        else:
            client = Client.objects.using(DB_NAME).get(id=18)
        if survey_type:
            projects = Project.objects.filter(survey_type=survey_type)
        else:
            return HttpResponse(status=404)
        project_locations = {}
        for project in projects:
            project_locations[project] = project.location_set.all()
        template = 'aecon/radar-download.html'
        return render(request, template,
                      {"client":      client, "projects": projects, "project_locations": project_locations,
                       "survey_type": survey_type})


def data_download_thread(project_location_list, user, survey_type_flag, root_url):
    if survey_type_flag in [9, 10]:
        class_lst = {43107: 1, 43108: 2, 43109: 3, 43110: 4, 43111: 5, 43112: 6, 43117: 7, 43118: 8, 43119: 9,
                     43120: 10, 43121: 11, 43122: 12}
    elif survey_type_flag == 12:
        class_lst = {43137: 1, 43138: 2, 43139: 3, 43140: 4, 43141: 5}

    survey_type_flag_dict = {9: "ATC", 10: "Link", 12: "Radar"}
    survey_type = survey_type_flag_dict[survey_type_flag]

    dir_lst = {1: 'N', 2: 'S', 9: 'E', 10: 'W'}

    project_location_dict = {}
    for project_location in project_location_list:
        project_id, location_id = str(project_location).split('_')
        if project_id not in project_location_dict.keys():
            project_location_dict[project_id] = []
        project_location_dict[project_id].append(location_id)

    print('Prepared request!')
    excel_streams = {}

    for project_id, location_list in project_location_dict.items():
        for location_id in location_list:
            observations = None
            if survey_type_flag == 9 or survey_type_flag == 12:
                observations = AssociatedObservation.objects.using(DB_NAME).filter(
                        location_id=location_id, project_id=project_id).values('date', 'value', 'obsClass', 'direction')
            elif survey_type_flag == 10:
                observations = LINKObservation.objects.using(DB_NAME).filter(
                        location_id=location_id, project_id=project_id).values('date', 'value', 'obsClass', 'direction',
                                                                               'weather_code', 'day_no', 'time_15min',
                                                                               'time_1hr')

            db_df = pd.DataFrame.from_records(observations.all())

            obs_classes = sorted(db_df.obsClass.unique())
            obs_directions = sorted(db_df.direction.unique())

            project = Project.objects.using(DB_NAME).get(id=project_id)
            location = Location.objects.using(DB_NAME).get(id=location_id)
            project_location = ProjectLocations.objects.using(DB_NAME).get(
                    project_id=project_id, location_id=location_id)

            DirectionList = location.directions.filter(
                    id__in=obs_directions).values("direction_id", "order")
            ClassList = location.classes.filter(
                    id__in=obs_classes).values("obsClass_id")

            classes_mapping = {}
            directions_mapping = {}

            for i in range(len(obs_directions)):
                try:
                    directions_mapping[obs_directions[i]] = dir_lst[DirectionList[i]
                    ["direction_id"]] + str(DirectionList[i]["order"])
                except:
                    print("wrong observation direction value")
            print("ClassList:- ", ClassList)
            for i in range(len(obs_classes)):
                try:
                    classes_mapping[obs_classes[i]
                    ] = class_lst[ClassList[i]["obsClass_id"]]
                except Exception as e:
                    print("wrong observation class value")
                    print(e)
            df_to_download = pd.DataFrame()
            if survey_type_flag == 9 or survey_type_flag == 12:
                df_to_download = pd.DataFrame(
                        columns=["Date", "Time", "Direction", "Speed", "Class"])
            elif survey_type_flag == 10:
                df_to_download = pd.DataFrame(
                        columns=['Date', 'Time', 'Direction', 'Weather_Code', 'Class', 'Volume', 'Day No', 'Time 15min',
                                 'Time 1hr'])

            try:
                df_to_download['Date'] = pd.to_datetime(db_df['date']).dt.date
                df_to_download['Time'] = pd.to_datetime(db_df['date']).dt.time
            except:
                print("dt.time is not working")

            print(survey_type_flag)

            if survey_type_flag == 10:
                df_to_download['Weather_Code'] = db_df['weather_code']
                df_to_download['Day No'] = db_df['day_no']
                df_to_download['Time 15min'] = db_df['time_15min']
                df_to_download['Time 1hr'] = db_df['time_1hr']
                df_to_download['Volume'] = db_df['value']
                df_to_download['Class'] = db_df['obsClass']
            elif survey_type_flag == 9 or survey_type_flag == 12:
                df_to_download['Speed'] = db_df['value']

            for i in range(len(df_to_download)):
                df_to_download.loc[i,
                "Direction"] = directions_mapping[db_df.loc[i, "direction"]]
                df_to_download.loc[i,
                "Class"] = classes_mapping[db_df.loc[i, "obsClass"]]

            def nearest_15_min(time):
                minutes = math.floor(time.minute / 15) * 15
                return time.replace(minute=minutes, second=0)

            if survey_type_flag == 9 or survey_type_flag == 12:
                df_to_download['Time 15min'] = df_to_download['Time'].apply(
                        lambda x: nearest_15_min(x))
                df_to_download['Time 1hr'] = df_to_download['Time'].apply(
                        lambda x: x.replace(minute=0, second=0))
                df_to_download['Speed Bin'] = df_to_download['Speed'].apply(
                        lambda x: math.floor(x / 5) * 5)
                df_to_download['Day No'] = df_to_download["Date"].apply(
                        lambda x: x.weekday() + 1)

            print('Prepared download dataframe')

            wb = Workbook()
            raw_data_sheet = wb.active
            for i, col in enumerate(df_to_download.columns, start=1):
                raw_data_sheet.cell(row=11, column=i, value=col)
            start_cell_row = 12
            start_cell_col = 1

            for i in range(df_to_download.shape[0]):
                for j in range(df_to_download.shape[1]):
                    raw_data_sheet.cell(row=start_cell_row + i, column=start_cell_col +
                                                                       j, value=df_to_download.iloc[i, j])

            primary_direction = dir_lst[DirectionList.get(order=0)[
                'direction_id']] + '0'
            secondary_direction = dir_lst[DirectionList.get(order=1)[
                'direction_id']] + '1'

            print('Written sheet 1 data')

            dashboard = wb.active
            dashboard.column_dimensions['A'].width = 12

            dashboard.merge_cells('B1:F1')
            dashboard.merge_cells('B2:F2')
            dashboard.merge_cells('B3:F3')
            dashboard.merge_cells('B4:F4')
            dashboard.merge_cells('B5:F5')
            dashboard.merge_cells('B6:F6')
            dashboard.merge_cells('B7:F7')
            dashboard.merge_cells('B8:F8')

            project_no = project.project_no
            location_name = location.name

            start_date_obj = project_location.startDate
            day = start_date_obj.strftime('%d')
            day_name = start_date_obj.strftime('%A')
            month_name = start_date_obj.strftime('%B')
            year = start_date_obj.strftime('%Y')
            start_date = '%s %s %s %s' % (day_name, str(day), month_name, year)

            end_date_obj = project_location.endDate
            day = end_date_obj.strftime('%d')
            day_name = end_date_obj.strftime('%A')
            month_name = end_date_obj.strftime('%B')
            year = end_date_obj.strftime('%Y')
            end_date = '%s %s %s %s' % (day_name, str(day), month_name, year)

            speed_limit = project_location.speed_limit if survey_type_flag == 9 or survey_type_flag == 12 else "-"
            lat = location.lat
            lon = location.lon

            print("Speed Limit: ", speed_limit)

            dashboard['A1'] = 'Client'
            client_cell = dashboard.cell(row=1, column=2)
            client_cell.value = 'AECON'
            client_cell.alignment = Alignment(
                    horizontal='right', vertical='center')

            dashboard['A2'] = 'Project'
            project_cell = dashboard.cell(row=2, column=2)
            project_cell.value = '%s - %s Data' % (
                    project_no, survey_type)
            project_cell.alignment = Alignment(
                    horizontal='right', vertical='center')

            dashboard['A3'] = 'Site'
            location_cell = dashboard.cell(row=3, column=2)
            location_cell.value = location_name
            location_cell.alignment = Alignment(
                    horizontal='right', vertical='center')

            dashboard['A4'] = 'Start Date'
            startDt_cell = dashboard.cell(row=4, column=2)
            startDt_cell.value = start_date
            startDt_cell.alignment = Alignment(
                    horizontal='right', vertical='center')

            dashboard['A5'] = 'End Date'
            endDt_cell = dashboard.cell(row=5, column=2)
            endDt_cell.value = end_date
            endDt_cell.alignment = Alignment(
                    horizontal='right', vertical='center')

            dashboard['A6'] = 'Lat/Long'
            lat_lon_cell = dashboard.cell(row=6, column=2)
            lat_lon_cell.value = '%s, %s' % (lat, lon)
            lat_lon_cell.alignment = Alignment(
                    horizontal='right', vertical='center')

            dashboard['A7'] = 'Primary Dir'
            primary_direction_cell = dashboard.cell(row=7, column=2)
            primary_direction_cell.value = primary_direction
            primary_direction_cell.alignment = Alignment(
                    horizontal='right', vertical='center')

            dashboard['A8'] = 'Secondary Dir'
            secondary_direction_cell = dashboard.cell(row=8, column=2)
            secondary_direction_cell.value = secondary_direction
            secondary_direction_cell.alignment = Alignment(
                    horizontal='right', vertical='center')

            if survey_type_flag != 10:
                dashboard.merge_cells('B9:F9')
                dashboard['A9'] = 'Speed Limit'
                speed_limit_cell = dashboard.cell(row=9, column=2)
                speed_limit_cell.value = speed_limit
                speed_limit_cell.alignment = Alignment(
                        horizontal='right', vertical='center')

            print('Written sheet 1 dashboard')
            # creating excel workbooks streams
            with NamedTemporaryFile(delete=False) as filetmp:
                wb.save(filetmp)
                filetmp.seek(0)
                filestream = filetmp.read()
                filename = '%s_%s.xlsx' % (
                        str(project_no), str(location_name))
                excel_streams[filename] = filestream
            print('Created excel file')

    token = generate_salt(16) + str(datetime.datetime.now().date())
    zip_path = os.path.join(EXCEL_ROOT, token)

    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as archive:
        for name, stream in excel_streams.items():
            archive.writestr(zinfo_or_arcname=name, data=stream)

    url = f'{root_url}downloadFile?token={token}&name=Tracsis%20{survey_type}%20{str(datetime.datetime.now().date())}&extension=zip'
    email = str(user.email)
    content = "<p>Dear " + str(user.first_name) + ",</p>"
    content += f"<p>The {survey_type} Data you requested from AECON is ready. Please note that, <a href='{url}'> link to the data, this link is available for 15 days only.</a></p>"
    content += "<p>Please don’t hesitate to contact us, If you have any issue.</p>"
    content += "<p>Kind Regards,"
    content += "<br>AECON Team</p>"
    content += "<p>ttd.admin@tracsis.com</p>"
    sendMail(email, 'Data Download is Ready', content)
    print('Email Sent')
    return


@login_required(login_url=LOGIN_URL)
def get_jtc_data(request):
    try:
        params = request.POST.dict()
        locationid = request.POST["location_id"].split(',')
        projectid = request.POST["project_id"].split(',')
        # set_cookie("Location", ids, max_age = None, expires = None)
        data = Jtc_Data.objects.using(DB_NAME).filter(location_id__in=locationid, project_id__in=projectid,
                                                      count__gt=0).values_list(
                'location_id', 'project_id', 'origin_arm_id', 'destination_arm_id', 'obsClass_id', 'start_time',
                'end_time', 'count', 'pcu')

        if len(data):
            df = pd.DataFrame.from_records(data, columns=[
                    'location_id', 'project_id', 'origin_arm_id', 'destination_arm_id', 'obsClass_id', 'start_time',
                    'end_time', 'count', 'pcu'])

            # Combine date with time to create a full datetime object
            df['start_time'] = pd.to_datetime(
                    df['start_time'].apply(lambda x: f'1970-01-01 {x}'))
            df['end_time'] = pd.to_datetime(
                    df['end_time'].apply(lambda x: f'1970-01-01 {x}'))

            # Group by 'origin_arm_id' and 'destination_arm_id'

            df.to_csv('teset_data.csv')
            grouped = df.groupby(
                    ['origin_arm_id', 'destination_arm_id', 'location_id'])
            result = []
            for key, group_main in grouped:
                obsClass_group = group_main.groupby(
                        pd.Grouper(key='start_time', freq='15T')).sum()
                arms_dict = {'origin_arm':      int(key[0]), 'destination_arm': int(
                        key[1]), 'location_id': int(key[2]), 'value': []}

                for index, row in obsClass_group.iterrows():
                    obs_class_data = group_main[(group_main['start_time'] >= index) & (
                            group_main['start_time'] < index + pd.Timedelta(minutes=15))]

                    obs_class_list = [{'id': class_row['obsClass_id'], 'count': class_row['count']}
                                      for _, class_row in obs_class_data.iterrows()]

                    arms_dict['value'].append({
                            'Start_time': index.strftime('%H:%M'),
                            'End_time':   (index + pd.Timedelta(minutes=14, seconds=59)).strftime('%H:%M'),
                            'class':      obs_class_list
                    })

                result.append(arms_dict)

            response = {'status': 'success', 'data': result}
        else:
            response = {'status': 'success', 'data': []}
    except Exception as e:
        response = {'status': 'error', 'message': str(e)}

    return JsonResponse(response)


@login_required(login_url=LOGIN_URL)
def add_update_jtc_Observation(request):
    try:
        loc_id = request.POST.get("location_id")
        project_id = request.POST.get("project_id")
        Observation = request.POST.get("Observation")
        print('loc_id', loc_id)
        print(Observation)
        location_instance = Location.objects.get(id=loc_id)
        print(location_instance)
        ObservationMessages, created = Messages.objects.using(DB_NAME).update_or_create(location=location_instance,
                                                                                        defaults={
                                                                                                'location': location_instance,
                                                                                                'text':     Observation,
                                                                                                'user':     request.user,
                                                                                                'date':     datetime.datetime.now().date()})
        response = {'status': 'success', 'data': created}
    except Exception as e:
        response = {'status': 'error', 'message': str(e)}
    return JsonResponse(response)


@login_required(login_url=LOGIN_URL)
def fetch_jtc_Observation(request):
    try:
        loc_id = request.POST.get("location_id")
        location_instance = Location.objects.get(id=loc_id)
        observations = Messages.objects.using(DB_NAME).filter(
                location=location_instance).values('text')
        response = list(observations)
    except Exception as e:
        response = {'status': 'error', 'message': str(e)}
    return JsonResponse(response, safe=False)


@login_required(login_url=LOGIN_URL)
def addUpdateObservation(request):
    try:
        loc_id = request.POST.get("location_id")
        project_id = request.POST.get("project_id", None)
        Observation = request.POST.get("Observation")

        location_instance = Location.objects.get(id=loc_id)
        ObservationMessages, created = Messages.objects.update_or_create(location=location_instance, project=project_id,
                                                                         defaults={
                                                                                 'location': location_instance,
                                                                                 'text':     Observation,
                                                                                 'user':     request.user,
                                                                                 'date':     datetime.datetime.now().date(),
                                                                                 'project':  project_id})
        response = {'status': 'success', 'data': created}
    except Exception as e:
        response = {'status': 'error', 'message': str(e)}
    return JsonResponse(response)


@login_required(login_url=LOGIN_URL)
def fetchObservation(request):
    try:
        loc_id = request.POST.get("location_id")
        project_id = request.POST.get("project_id", None)
        location_instance = Location.objects.get(id=loc_id)
        observations = Messages.objects.filter(location=location_instance, project=project_id).values("text")
        response = list(observations)
    except Exception as e:
        response = {'status': 'error', 'message': str(e)}
    return JsonResponse(response, safe=False)
